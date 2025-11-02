from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Dict, List

import openpyxl

from core.conversion_engine import ConversionEngine

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "homepage" / "database" / "app.db"
INPUT_FILE = Path("tests/input/sample_invoice2.xlsx")
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

supplier_keys = {
    "supplier_name": "company_name",
    "supplier_business_number": "business_number",
    "supplier_representative": "representative_name",
    "supplier_address": "address",
    "supplier_business_type": "business_type",
    "supplier_business_category": "business_category",
    "supplier_email": "email",
}

def fetch_gold_users() -> List[Dict[str, str]]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT username, email, company_name, business_number, representative_name, address, phone, business_type, business_category FROM users WHERE plan_type = ? ORDER BY username",
        ("gold-vip",),
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]

def build_supplier_info(user: Dict[str, str]) -> Dict[str, str]:
    info: Dict[str, str] = {}
    for target_key, source_key in supplier_keys.items():
        info[target_key] = user.get(source_key, "") or ""
    return info

def inspect_supplier_cells(path: Path) -> Dict[str, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = "?묒??낅줈?쒖뼇?? if "?묒??낅줈?쒖뼇?? in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    row = 7
    cells = {
        "business_number": ws[f"C{row}"].value,
        "supplier_name": ws[f"E{row}"].value,
        "representative": ws[f"F{row}"].value,
        "address": ws[f"G{row}"].value,
        "business_type": ws[f"H{row}"].value,
        "business_category": ws[f"I{row}"].value,
        "email": ws[f"J{row}"].value,
    }
    wb.close()
    return cells

def main() -> None:
    gold_users = fetch_gold_users()
    if not gold_users:
        print("[ERROR] 怨⑤뱶 ?뚯썝??李얠쓣 ???놁뒿?덈떎.")
        return

    print(f"[INFO] 怨⑤뱶 ?뚯썝 {len(gold_users)}紐?蹂???뚯뒪???쒖옉")
    engine = ConversionEngine()
    for user in gold_users:
        supplier_info = build_supplier_info(user)
        file_name = f"gold_{user['username']}"
        result = engine.convert_file(
            uploaded_file_path=str(INPUT_FILE),
            supplier_info=supplier_info,
            template_id="hometax_official",
            industry_type="delivery",
            issue_date=None,
            file_name=file_name,
            user_info={"user_id": user['username']},
        )

        if not result.get("success"):
            print(f"[FAIL] {user['username']} 蹂???ㅽ뙣: {result.get('error_message')}")
            continue

        first_file = Path(result['files'][0])
        cells = inspect_supplier_cells(first_file)
        print(f"[OK] {user['username']} ??{first_file.name}")
        print(
            "   怨듦툒???뺣낫:",
            cells["business_number"],
            cells["supplier_name"],
            cells["representative"],
        )
    print("[INFO] ?뚯뒪???꾨즺")


if __name__ == "__main__":
    main()