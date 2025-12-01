"""
세무 리포트 API 모듈
상용화 준비: 부가세 신고를 위한 전용 통계 시스템
"""

from flask import Blueprint, request, jsonify, send_file, session
from core.db import get_conn_optimized as get_conn
from core.responses import success, error
import sqlite3
import logging
from datetime import datetime, timedelta
import pandas as pd
import io
from ..utils.auth import ensure_admin_for_json

logger = logging.getLogger(__name__)

admin_tax_bp = Blueprint('admin_tax', __name__, url_prefix='/admin/api/tax')


@admin_tax_bp.route('/stats', methods=['GET'])
def get_tax_stats():
    """
    세무 통계 조회 API
    
    Query Parameters:
        type: daily, monthly, quarterly, yearly (기본값: monthly)
        year: 연도 (기본값: 현재 연도)
        month: 월 (1-12, 기본값: 현재 월)
        start_date: 시작일 (YYYY-MM-DD 형식)
        end_date: 종료일 (YYYY-MM-DD 형식)
    
    Response:
        {
            "success": true,
            "data": {
                "total_revenue": 0,      # 총 매출 (부가세 포함)
                "total_supply": 0,      # 총 공급가액
                "total_vat": 0,          # 총 부가세 (납부예정액)
                "order_count": 0,       # 주문 건수
                "chart_data": [         # 차트용 데이터
                    {
                        "date": "2024-01",
                        "revenue": 0,
                        "supply": 0,
                        "vat": 0
                    }
                ]
            }
        }
    """
    # 관리자 인증 확인
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    # 파라미터 파싱
    report_type = request.args.get('type', 'monthly')  # daily, monthly, quarterly, yearly
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 날짜 범위 계산
    now = datetime.now()
    if not year:
        year = now.year
    if not month:
        month = now.month
    
    try:
        with get_conn() as conn:
            conn.row_factory = sqlite3.Row
            
            # 날짜 범위 설정
            if start_date and end_date:
                # 직접 지정된 날짜 범위 사용
                date_filter = "DATE(o.created_at) BETWEEN ? AND ?"
                date_params = (start_date, end_date)
                date_format = "%Y-%m-%d"
            elif report_type == 'daily':
                # 일별: 지정된 날짜 또는 오늘
                target_date = f"{year}-{month:02d}-{now.day:02d}" if month else now.strftime("%Y-%m-%d")
                date_filter = "DATE(o.created_at) = ?"
                date_params = (target_date,)
                date_format = "%Y-%m-%d"
            elif report_type == 'monthly':
                # 월별: 지정된 연도-월
                date_filter = "strftime('%Y-%m', o.created_at) = ?"
                date_params = (f"{year}-{month:02d}",)
                date_format = "%Y-%m"
            elif report_type == 'quarterly':
                # 분기별: 지정된 연도의 분기
                quarter = (month - 1) // 3 + 1 if month else (now.month - 1) // 3 + 1
                start_month = (quarter - 1) * 3 + 1
                end_month = quarter * 3
                date_filter = "strftime('%Y-%m', o.created_at) BETWEEN ? AND ?"
                date_params = (f"{year}-{start_month:02d}", f"{year}-{end_month:02d}")
                date_format = "%Y-%m"
            elif report_type == 'yearly':
                # 년별: 지정된 연도
                date_filter = "strftime('%Y', o.created_at) = ?"
                date_params = (str(year),)
                date_format = "%Y"
            else:
                # 기본값: 이번 달
                date_filter = "strftime('%Y-%m', o.created_at) = ?"
                date_params = (now.strftime("%Y-%m"),)
                date_format = "%Y-%m"
            
            # 주문 데이터 조회 (paid + cancelled 모두 포함, payment_history와 조인하여 pg_provider도 가져옴)
            orders_query = f"""
                SELECT 
                    o.id,
                    o.merchant_uid,
                    o.status,
                    o.amount,
                    o.supply_price,
                    o.vat,
                    o.created_at,
                    o.product_name,
                    o.payment_method,
                    ph.pg_provider,
                    u.username
                FROM orders o
                LEFT JOIN users u ON o.user_id = u.id
                LEFT JOIN payment_history ph ON o.merchant_uid = ph.order_id
                WHERE o.status IN ('paid', 'cancelled')
                  AND {date_filter}
                ORDER BY o.created_at DESC
            """
            
            orders_rows = conn.execute(orders_query, date_params).fetchall()
            
            # Pandas DataFrame으로 변환
            if orders_rows:
                orders_data = [dict(row) for row in orders_rows]
                df = pd.DataFrame(orders_data)
            else:
                df = pd.DataFrame(columns=['id', 'merchant_uid', 'status', 'amount', 'supply_price', 'vat', 'created_at', 'product_name', 'payment_method', 'pg_provider', 'username'])
            
            # 집계 계산 (Pandas 활용)
            total_revenue = int(df['amount'].sum()) if not df.empty else 0  # 총 거래액 (paid + cancelled)
            refunds = int(df[df['status'] == 'cancelled']['amount'].sum()) if not df.empty else 0  # 환불액
            net_revenue = int(df[df['status'] == 'paid']['amount'].sum()) if not df.empty else 0  # 순 매출
            total_vat = int(df[df['status'] == 'paid']['vat'].sum()) if not df.empty else 0  # 부가세 (순 매출 기준)
            total_supply = int(df[df['status'] == 'paid']['supply_price'].sum()) if not df.empty else 0  # 공급가액 (순 매출 기준)
            order_count = len(df[df['status'] == 'paid']) if not df.empty else 0  # 주문 건수 (paid만)
            
            # 결제 수단별 통계 계산 (순 매출 기준) - 3단 분류: 카드/현금(증빙)/기타(무증빙)
            paid_df = df[df['status'] == 'paid'] if not df.empty else pd.DataFrame()
            if not paid_df.empty:
                # payment_method 우선, 없으면 pg_provider 사용하여 3단 분류
                def categorize_payment_method(row):
                    # payment_method 우선 확인
                    method = row.get('payment_method') if 'payment_method' in row else None
                    pg_provider = row.get('pg_provider') if 'pg_provider' in row else None
                    
                    # payment_method가 없거나 NULL이면 pg_provider 사용
                    if pd.isna(method) or method == '' or method is None:
                        method = pg_provider
                    
                    # 둘 다 없으면 'other'
                    if pd.isna(method) or method == '' or method is None:
                        return 'other'
                    
                    method_lower = str(method).lower()
                    
                    # 카드: 'card' 포함
                    if method_lower == 'card' or 'card' in method_lower:
                        return 'card'
                    # 현금(증빙): 'trans', 'vbank' 포함
                    elif method_lower in ['trans', 'vbank'] or 'trans' in method_lower or 'vbank' in method_lower:
                        return 'cash'
                    # 기타(무증빙): 'manual', 'virtual', 'test_virtual', 'unknown', NULL 등
                    else:
                        return 'other'
                
                # 결제 수단 카테고리 추가
                paid_df['payment_category'] = paid_df.apply(categorize_payment_method, axis=1)
                
                # 3단 분류 집계
                card_total = int(paid_df[paid_df['payment_category'] == 'card']['amount'].sum()) if 'payment_category' in paid_df.columns else 0
                cash_total = int(paid_df[paid_df['payment_category'] == 'cash']['amount'].sum()) if 'payment_category' in paid_df.columns else 0
                other_total = int(paid_df[paid_df['payment_category'] == 'other']['amount'].sum()) if 'payment_category' in paid_df.columns else 0
            else:
                card_total = 0
                cash_total = 0
                other_total = 0
            
            # 차트 데이터 조회 (일별/월별 그룹화) - 순 매출 기준
            if report_type == 'daily':
                # 일별 데이터
                chart_query = f"""
                    SELECT 
                        DATE(o.created_at) as date,
                        COUNT(CASE WHEN o.status = 'paid' THEN 1 END) as order_count,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.amount ELSE 0 END), 0) as revenue,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.supply_price ELSE 0 END), 0) as supply,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.vat ELSE 0 END), 0) as vat
                    FROM orders o
                    WHERE o.status IN ('paid', 'cancelled')
                      AND {date_filter}
                    GROUP BY DATE(o.created_at)
                    ORDER BY date ASC
                """
            elif report_type == 'monthly':
                # 월별 데이터
                chart_query = f"""
                    SELECT 
                        strftime('%Y-%m', o.created_at) as date,
                        COUNT(CASE WHEN o.status = 'paid' THEN 1 END) as order_count,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.amount ELSE 0 END), 0) as revenue,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.supply_price ELSE 0 END), 0) as supply,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.vat ELSE 0 END), 0) as vat
                    FROM orders o
                    WHERE o.status IN ('paid', 'cancelled')
                      AND {date_filter}
                    GROUP BY strftime('%Y-%m', o.created_at)
                    ORDER BY date ASC
                """
            elif report_type == 'quarterly':
                # 분기별 데이터
                chart_query = f"""
                    SELECT 
                        strftime('%Y-Q', o.created_at) || CAST((CAST(strftime('%m', o.created_at) AS INTEGER) - 1) / 3 + 1 AS TEXT) as date,
                        COUNT(CASE WHEN o.status = 'paid' THEN 1 END) as order_count,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.amount ELSE 0 END), 0) as revenue,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.supply_price ELSE 0 END), 0) as supply,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.vat ELSE 0 END), 0) as vat
                    FROM orders o
                    WHERE o.status IN ('paid', 'cancelled')
                      AND {date_filter}
                    GROUP BY strftime('%Y', o.created_at), (CAST(strftime('%m', o.created_at) AS INTEGER) - 1) / 3 + 1
                    ORDER BY date ASC
                """
            else:  # yearly
                # 년별 데이터
                chart_query = f"""
                    SELECT 
                        strftime('%Y', o.created_at) as date,
                        COUNT(CASE WHEN o.status = 'paid' THEN 1 END) as order_count,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.amount ELSE 0 END), 0) as revenue,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.supply_price ELSE 0 END), 0) as supply,
                        COALESCE(SUM(CASE WHEN o.status = 'paid' THEN o.vat ELSE 0 END), 0) as vat
                    FROM orders o
                    WHERE o.status IN ('paid', 'cancelled')
                      AND {date_filter}
                    GROUP BY strftime('%Y', o.created_at)
                    ORDER BY date ASC
                """
            
            chart_rows = conn.execute(chart_query, date_params).fetchall()
            chart_data = [
                {
                    'date': row['date'],
                    'revenue': row['revenue'] or 0,
                    'supply': row['supply'] or 0,
                    'vat': row['vat'] or 0,
                    'order_count': row['order_count'] or 0
                }
                for row in chart_rows
            ]
            
            # 상세 거래 내역 (최근 거래 내역)
            recent_transactions = df.head(100).to_dict('records') if not df.empty else []
            
            return success('세무 통계 조회 성공', data={
                'total_revenue': total_revenue,  # 총 거래액 (paid + cancelled)
                'refunds': refunds,  # 환불액
                'net_revenue': net_revenue,  # 순 매출
                'total_vat': total_vat,  # 부가세
                'total_supply': total_supply,  # 공급가액
                'order_count': order_count,  # 주문 건수
                'card_total': card_total,  # 카드 결제 총액
                'cash_total': cash_total,  # 현금(증빙) 결제 총액
                'other_total': other_total,  # 기타(무증빙) 결제 총액
                'chart_data': chart_data,
                'recent_transactions': recent_transactions,  # 상세 거래 내역
                'report_type': report_type,
                'date_range': {
                    'start': start_date or (f"{year}-{month:02d}-01" if month else f"{year}-01-01"),
                    'end': end_date or (f"{year}-{month:02d}-{now.day:02d}" if month else f"{year}-12-31")
                }
            })
            
    except Exception as e:
        logger.error(f"세무 통계 조회 실패: {str(e)}")
        return error(f'세무 통계 조회 중 오류가 발생했습니다: {str(e)}', status=500)


@admin_tax_bp.route('/download', methods=['GET'])
def download_tax_report():
    """
    세무 리포트 엑셀 다운로드 API
    
    Query Parameters:
        type: daily, monthly, quarterly, yearly (기본값: monthly)
        year: 연도
        month: 월 (1-12)
        start_date: 시작일 (YYYY-MM-DD)
        end_date: 종료일 (YYYY-MM-DD)
    
    Response:
        Excel 파일 (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
        컬럼: 일자, 주문번호, 상품명, 공급가, 부가세, 합계
    """
    # 관리자 인증 확인
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    # 파라미터 파싱
    report_type = request.args.get('type', 'monthly')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 날짜 범위 계산
    now = datetime.now()
    if not year:
        year = now.year
    if not month:
        month = now.month
    
    try:
        with get_conn() as conn:
            conn.row_factory = sqlite3.Row
            
            # 날짜 범위 설정
            if start_date and end_date:
                date_filter = "DATE(o.created_at) BETWEEN ? AND ?"
                date_params = (start_date, end_date)
            elif report_type == 'daily':
                target_date = f"{year}-{month:02d}-{now.day:02d}" if month else now.strftime("%Y-%m-%d")
                date_filter = "DATE(o.created_at) = ?"
                date_params = (target_date,)
            elif report_type == 'monthly':
                date_filter = "strftime('%Y-%m', o.created_at) = ?"
                date_params = (f"{year}-{month:02d}",)
            elif report_type == 'quarterly':
                quarter = (month - 1) // 3 + 1 if month else (now.month - 1) // 3 + 1
                start_month = (quarter - 1) * 3 + 1
                end_month = quarter * 3
                date_filter = "strftime('%Y-%m', o.created_at) BETWEEN ? AND ?"
                date_params = (f"{year}-{start_month:02d}", f"{year}-{end_month:02d}")
            elif report_type == 'yearly':
                date_filter = "strftime('%Y', o.created_at) = ?"
                date_params = (str(year),)
            else:
                date_filter = "strftime('%Y-%m', o.created_at) = ?"
                date_params = (now.strftime("%Y-%m"),)
            
            # 주문 내역 조회 (users 테이블 조인하여 고객명 포함)
            orders = conn.execute(
                f"""
                SELECT 
                    DATE(o.created_at) as 일자,
                    o.merchant_uid as 주문번호,
                    COALESCE(u.username, '알 수 없음') as 고객명,
                    o.product_name as 상품명,
                    o.supply_price as 공급가액,
                    o.vat as 세액,
                    o.amount as 합계금액,
                    CASE 
                        WHEN o.status = 'paid' THEN '정상'
                        WHEN o.status = 'cancelled' THEN '취소'
                        ELSE o.status
                    END as 상태
                FROM orders o
                LEFT JOIN users u ON o.user_id = u.id
                WHERE o.status IN ('paid', 'cancelled')
                  AND {date_filter}
                ORDER BY o.created_at ASC
                """,
                date_params
            ).fetchall()
            
            # DataFrame 생성
            df = pd.DataFrame([dict(row) for row in orders])
            
            if df.empty:
                # 빈 데이터프레임인 경우 기본 구조만 생성
                df = pd.DataFrame(columns=['일자', '주문번호', '고객명', '상품명', '공급가액', '세액', '합계금액', '상태'])
            
            # 엑셀 파일 생성 (메모리 버퍼)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='세무 리포트', index=False)
                
                # openpyxl 스타일링
                from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
                from openpyxl.utils import get_column_letter
                
                worksheet = writer.sheets['세무 리포트']
                
                # 헤더 행 스타일링 (1행)
                header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                header_font = Font(bold=True, size=11)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 금액 컬럼 포맷팅 및 정렬
                amount_columns = ['공급가액', '세액', '합계금액']
                amount_column_indices = []
                for idx, col in enumerate(df.columns, start=1):
                    if col in amount_columns:
                        amount_column_indices.append(idx)
                        # 숫자 포맷 적용 (3자리 콤마)
                        for row in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row, column=idx)
                            if cell.value is not None:
                                try:
                                    cell.number_format = '#,##0'
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                                except:
                                    pass
                
                # 컬럼 너비 자동 조정
                for idx, col in enumerate(df.columns, start=1):
                    max_length = 0
                    # 헤더 길이 확인
                    max_length = max(max_length, len(str(col)))
                    # 데이터 길이 확인
                    if not df.empty:
                        col_data = df[col].astype(str)
                        max_length = max(max_length, col_data.map(len).max())
                    # 컬럼 너비 설정 (최소 10, 최대 50)
                    worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 10), 50)
            
            output.seek(0)
            
            # 파일명 생성
            if start_date and end_date:
                filename = f"세무리포트_{start_date}_{end_date}.xlsx"
            elif report_type == 'daily':
                filename = f"세무리포트_일별_{year}{month:02d}{now.day:02d}.xlsx"
            elif report_type == 'monthly':
                filename = f"세무리포트_월별_{year}{month:02d}.xlsx"
            elif report_type == 'quarterly':
                quarter = (month - 1) // 3 + 1 if month else (now.month - 1) // 3 + 1
                filename = f"세무리포트_분기별_{year}Q{quarter}.xlsx"
            else:  # yearly
                filename = f"세무리포트_년별_{year}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
    except Exception as e:
        logger.error(f"세무 리포트 다운로드 실패: {str(e)}")
        return error(f'세무 리포트 다운로드 중 오류가 발생했습니다: {str(e)}', status=500)

