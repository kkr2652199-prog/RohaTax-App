from flask import Blueprint, render_template, session, redirect, url_for, jsonify, request
from urllib.parse import quote
import os
from core.db import get_conn_optimized as get_conn
from core.responses import success, error
# from core.template_manager import template_manager
from core.data_bus import validate_convert_start, normalize_convert_start, SCHEMA_VERSION
from core.absolute_guidelines import absolute_guidelines
from core.file_validator import file_validator
from core.notification_system import notification_system
from core.security import generate_csrf_token
from core.conversion_engine import ConversionEngine
from core.subscription_utils import get_user_subscription, is_unlimited_user, update_plan_price_and_tokens
from core.file_upload_helper import save_uploaded_file, cleanup_temp_file, calculate_template_count
from core.token_deduction_processor import TokenDeductionProcessor  # 연동 모듈 추가
from core.activity_service import record_activity  # 기록관 연동 모듈 추가
from datetime import datetime
import time
import logging
import sqlite3

from .utils.auth import (
    current_user_id,
    ensure_admin_for_json,
    ensure_login_for_json,
    is_authenticated,
)
from core.utils import row_value
from core.token_service import get_user_token_status

conversion_bp = Blueprint('conversion', __name__)

logger = logging.getLogger(__name__)

# 전역 변환 엔진 인스턴스 제거 (상태 격리를 위해)
# conversion_engine = ConversionEngine()  # 제거됨 - 요청별 인스턴스 생성으로 변경


def _calculate_template_count_precisely(uploaded_file, industry_type: str = 'delivery') -> int:
    """
    변환 전 템플릿 건수를 정밀하게 계산하는 함수
    
    실제 변환 로직과 동일한 방식으로 공급받는자 수를 계산하여
    필요한 토큰 수를 정확히 예측
    
    Args:
        uploaded_file: Flask uploaded file
        industry_type: 업종 타입 (기본값: 'delivery')
        
    Returns:
        int: 템플릿 건수 (공급받는자 수)
    """
    from core.file_parser import FileParser
    
    try:
        logger.info(f"템플릿 건수 정밀 계산 시작: 업종={industry_type}")
        
        # 임시 파일로 저장
        import tempfile
        temp_dir = tempfile.mkdtemp()
        temp_file_path = os.path.join(temp_dir, uploaded_file.filename)
        
        # 파일 포인터를 처음으로 이동 후 저장
        uploaded_file.seek(0)
        uploaded_file.save(temp_file_path)
        
        # 파일 파싱
        file_parser = FileParser()
        parsed_data = file_parser.parse_file(temp_file_path)
        
        logger.info(f"파싱 결과: {type(parsed_data)}")
        logger.info(f"파싱 결과 키: {list(parsed_data.keys()) if isinstance(parsed_data, dict) else 'Not a dict'}")
        
        if not parsed_data:
            logger.warning("파일 파싱 실패 또는 데이터 없음")
            return 0
        
        # 파싱된 데이터에서 행 수 계산
        # parse_file이 반환하는 구조: {'data_sections': {...}, 'total_rows': int}
        total_rows = parsed_data.get('total_rows', 0)
        
        if total_rows > 0:
            recipient_count = total_rows
            logger.info(f"파싱된 총 행 수: {recipient_count}")
        else:
            # 데이터 섹션에서 계산 시도
            data_sections = parsed_data.get('data_sections', {})
            data_section = data_sections.get('data_section', [])
            
            if isinstance(data_section, list):
                recipient_count = len(data_section)
            elif hasattr(data_section, '__len__'):
                recipient_count = len(data_section)
            else:
                # 기본값으로 50 반환 (임시)
                logger.warning("파싱된 데이터에서 행 수를 계산할 수 없음, 기본값 50 사용")
                recipient_count = 50
        
        # 임시 파일 정리
        try:
            os.remove(temp_file_path)
            os.rmdir(temp_dir)
        except Exception as cleanup_error:
            logger.warning(f"임시 파일 정리 실패: {cleanup_error}")
        
        logger.info(f"템플릿 건수 계산 완료: {recipient_count}건")
        
        return recipient_count
        
    except Exception as e:
        logger.error(f"템플릿 건수 계산 중 오류 발생: {str(e)}")
        # 실패 시 0 반환 (안전한 기본값)
        return 0


@conversion_bp.route('/conversion')
def conversion():
    # 로그인 확인
    if not is_authenticated():
        return render_template(
            'conversion.html',
            guest_mode=True,
            csrf_token=generate_csrf_token(),
        )

    return render_template(
        'conversion.html',
        guest_mode=False,
        csrf_token=generate_csrf_token(),
    )


@conversion_bp.route('/api/use-token', methods=['POST'])
def use_token():
    """변환 작업 시 토큰 사용 API"""
    user_id, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response
    
    data = request.get_json(silent=True) or {}
    tokens_to_use = int(data.get('tokens', 1))  # 기본 1토큰
    
    if tokens_to_use <= 0:
        return error('토큰 수량은 1 이상이어야 합니다', status=400)
    
    # 토큰 상태 조회 (중복 제거: token_service 사용)
    token_status = get_user_token_status(user_id)
    if token_status is None:
        return error('사용자를 찾을 수 없습니다', status=404)
    
    available_tokens = token_status['available_tokens']
    
    if available_tokens < tokens_to_use:
        return error(f'토큰이 부족합니다. 사용 가능: {available_tokens}개, 요청: {tokens_to_use}개', status=400)
    
    # 토큰 사용량 업데이트
    with get_conn() as conn:
        new_tokens_used = token_status['tokens_used'] + tokens_to_use
        conn.execute(
            "UPDATE users SET tokens_used = ? WHERE id = ?", 
            (new_tokens_used, user_id)
        )
        conn.commit()
    
    # 업데이트된 잔액 반환
    remaining_tokens = token_status['token_balance'] - new_tokens_used
    
    return success('토큰이 사용되었습니다', data={
        'tokens_used': tokens_to_use,
        'remaining_tokens': remaining_tokens,
        'total_granted': token_status['token_balance'],
        'total_used': new_tokens_used
    })


@conversion_bp.route('/api/token-status', methods=['GET'])
def token_status():
    """현재 토큰 상태 조회 API (캐싱 적용)"""
    user_id, guard_response = ensure_login_for_json()
    if guard_response is not None:
        logger.warning("로그인되지 않은 사용자")
        return guard_response

    logger.info(f"토큰상태 요청 - 세션 ID: {user_id}")
    logger.info(f"세션 전체: {dict(session)}")
    logger.info(f"요청 헤더: {dict(request.headers)}")

    # 토큰 잔액 조회 (중복 제거: token_service 사용)
    token_status = get_user_token_status(user_id)
    if token_status is None:
        logger.warning(f"사용자 ID {user_id}를 찾을 수 없음")
        return error('사용자를 찾을 수 없습니다', status=404)
    
    logger.info(f"토큰 상태 조회 성공: Balance={token_status['token_balance']}, Used={token_status['tokens_used']}, Available={token_status['available_tokens']}")
    
    return success('토큰 상태 조회 성공', data={
        'total_granted': token_status['token_balance'],
        'total_used': token_status['tokens_used'],
        'available_tokens': token_status['available_tokens']
    })


 


@conversion_bp.route('/api/user-info', methods=['GET'])
def user_info():
    """현재 로그인한 사용자 정보 조회 API"""
    user_id, guard_response = ensure_login_for_json()
    if guard_response is not None:
        logger.warning("로그인되지 않은 사용자")
        return guard_response

    logger.info(f"유저정보 요청 - 세션 ID: {user_id}")
    
    # 사용자 정보 조회 (기존 방식으로 복원)
    with get_conn() as conn:
        user = conn.execute(
            """SELECT id, username, email, company_name, business_number,
                      representative_name, phone, address, business_type, business_category,
                      plan_type, monthly_limit, used_count, is_active, is_admin,
                      token_balance, tokens_used, created_at
               FROM users WHERE id = ?""",
            (user_id,)
        ).fetchone()
        
        if not user:
            logger.warning(f"사용자 ID {user_id}를 찾을 수 없음")
            return error('사용자를 찾을 수 없습니다', status=404)
        
        logger.info(f"사용자 정보 조회 성공: {user['username']}")
        
        # 민감한 정보는 제외하고 필요한 정보만 반환
        safe_user_data = {
            'id': row_value(user, 'id'),
            'username': row_value(user, 'username', ''),
            'email': row_value(user, 'email', ''),
            'company_name': row_value(user, 'company_name', ''),
            'business_number': row_value(user, 'business_number', ''),
            'representative_name': row_value(user, 'representative_name', ''),
            'phone': row_value(user, 'phone', ''),
            'address': row_value(user, 'address', ''),
            'business_type': row_value(user, 'business_type', ''),
            'business_category': row_value(user, 'business_category', ''),
            'plan_type': row_value(user, 'plan_type', ''),
            'monthly_limit': row_value(user, 'monthly_limit', 0),
            'used_count': row_value(user, 'used_count', 0),
            'is_active': bool(row_value(user, 'is_active', 0)),
            'is_admin': bool(row_value(user, 'is_admin', 0)),
            'token_balance': row_value(user, 'token_balance', 0) or 0,
            'tokens_used': row_value(user, 'tokens_used', 0) or 0,
            'created_at': row_value(user, 'created_at', '')
        }
        
        return success('사용자 정보 조회 성공', data={'user': safe_user_data})


@conversion_bp.route('/api/templates', methods=['GET'])
def get_templates():
    """사용 가능한 템플릿 목록 조회 API"""
    _, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        templates = template_manager.get_available_templates()
        return success('템플릿 목록 조회 성공', data={'templates': templates})
    except Exception as e:
        return error(f'템플릿 목록 조회 실패: {str(e)}', status=500)


@conversion_bp.route('/api/templates/<template_id>', methods=['GET'])
def get_template_info(template_id):
    """특정 템플릿 정보 조회 API"""
    _, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        template_info = template_manager.get_template_info(template_id)
        if not template_info:
            return error('템플릿을 찾을 수 없습니다', status=404)
        
        return success('템플릿 정보 조회 성공', data={'template': template_info})
    except Exception as e:
        return error(f'템플릿 정보 조회 실패: {str(e)}', status=500)


@conversion_bp.route('/api/templates/<template_id>/validate', methods=['GET'])
def validate_template(template_id):
    """템플릿 파일 유효성 검사 API"""
    _, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        is_valid = template_manager.validate_template_file(template_id)
        template_path = template_manager.get_template_path(template_id)
        
        return success('템플릿 유효성 검사 완료', data={
            'template_id': template_id,
            'is_valid': is_valid,
            'file_path': template_path
        })
    except Exception as e:
        return error(f'템플릿 유효성 검사 실패: {str(e)}', status=500)


@conversion_bp.route('/api/templates/upload', methods=['POST'])
def upload_template():
    """템플릿 파일 업로드 API (관리자 전용)"""
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        # 업로드된 파일 확인
        if 'template_file' not in request.files:
            return error('템플릿 파일이 없습니다', status=400)
        
        template_file = request.files['template_file']
        if template_file.filename == '':
            return error('파일이 선택되지 않았습니다', status=400)
        
        # 파일 확장자 검증
        if not template_file.filename.lower().endswith(('.xlsx', '.xlsm', '.xls')):
            return error('Excel 파일만 업로드 가능합니다', status=400)
        
        # 템플릿 정보 파싱
        template_id = request.form.get('template_id')
        template_name = request.form.get('template_name')
        template_description = request.form.get('template_description')
        sheet_name = request.form.get('sheet_name', 'Sheet1')
        header_row = int(request.form.get('header_row', 1))
        
        if not all([template_id, template_name]):
            return error('템플릿 ID와 이름은 필수입니다', status=400)
        
        # 템플릿 디렉토리 생성
        template_dir = template_manager.create_template_directory(template_id)
        
        # 파일 저장
        filename = f"{template_id}_template.xlsx"
        file_path = os.path.join(template_dir, filename)
        template_file.save(file_path)
        
        # 템플릿 설정에 추가
        template_info = {
            "name": template_name,
            "description": template_description or f"{template_name} 템플릿",
            "file": f"{template_id}/{filename}",
            "sheet_name": sheet_name,
            "header_row": header_row,
            "fields": {}  # 나중에 필드 매핑 추가 가능
        }
        
        success_result = template_manager.add_template(template_id, template_info)
        
        if success_result:
            return success('템플릿 업로드 성공', data={
                'template_id': template_id,
                'file_path': file_path,
                'template_info': template_info
            })
        else:
            return error('템플릿 설정 저장 실패', status=500)
            
    except Exception as e:
        return error(f'템플릿 업로드 실패: {str(e)}', status=500)


@conversion_bp.route('/api/validate-template-data', methods=['POST'])
def validate_template_data():
    """템플릿 데이터 절대지침 검증 API"""
    user_id, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        data = request.get_json(silent=True) or {}
        
        # 절대지침 검증 수행
        success_result, errors = absolute_guidelines.validate_template_data(data)
        
        # 검증 결과 로그 기록
        absolute_guidelines.log_validation_result(
            data, success_result, errors, user_id
        )
        
        if success_result:
            return success('절대지침 검증 통과', data={
                'compliant': True,
                'errors': [],
                'guideline_version': absolute_guidelines.get_guideline_version()
            })
        else:
            return error('절대지침 검증 실패', data={
                'compliant': False,
                'errors': errors,
                'guideline_version': absolute_guidelines.get_guideline_version()
            }, status=422)
            
    except Exception as e:
        return error(f'검증 처리 실패: {str(e)}', status=500)


@conversion_bp.route('/api/guidelines/version', methods=['GET'])
def get_guidelines_version():
    """절대지침 버전 조회 API"""
    try:
        version = absolute_guidelines.get_guideline_version()
        return success('절대지침 버전 조회 성공', data={
            'version': version,
            'last_updated': '2025-10-01',
            'status': 'active'
        })
    except Exception as e:
               return error(f'버전 조회 실패: {str(e)}', status=500)


@conversion_bp.route('/api/security/status', methods=['GET'])
def get_security_status():
    """보안 시스템 상태 조회 API"""
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        # 파일 검증 시스템 상태
        validation_summary = file_validator.get_validation_summary()
        
        # 알림 시스템 상태
        notification_stats = notification_system.get_notification_stats()
        
        # 최근 알림 목록
        recent_notifications = notification_system.get_notifications(limit=10)
        
        return success('보안 시스템 상태 조회 성공', data={
            'file_validation': validation_summary,
            'notifications': notification_stats,
            'recent_notifications': recent_notifications,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return error(f'보안 시스템 상태 조회 실패: {str(e)}', status=500)


@conversion_bp.route('/api/security/notifications', methods=['GET'])
def get_notifications():
    """알림 목록 조회 API"""
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        category = request.args.get('category')
        priority = request.args.get('priority')
        unread_only = request.args.get('unread_only', 'false').lower() == 'true'
        limit = int(request.args.get('limit', 50))
        
        notifications = notification_system.get_notifications(
            category=category,
            priority=priority,
            unread_only=unread_only,
            limit=limit
        )
        
        return success('알림 목록 조회 성공', data={
            'notifications': notifications,
            'count': len(notifications)
        })
        
    except Exception as e:
        return error(f'알림 목록 조회 실패: {str(e)}', status=500)


@conversion_bp.route('/api/security/notifications/<int:notification_id>/read', methods=['POST'])
def mark_notification_read(notification_id):
    """알림을 읽음으로 표시 API"""
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        notification_system.mark_as_read(notification_id)
        return success('알림을 읽음으로 표시했습니다')
        
    except Exception as e:
        return error(f'알림 읽음 표시 실패: {str(e)}', status=500)


# 변환 시작: 파일 업로드 + 공급받는자 정보 추출 + 템플릿 기입
@conversion_bp.route('/api/convert/start', methods=['POST'])
def start_conversion():
    """변환 시작 API (파일 업로드 + 공급받는자 정보 추출 + 템플릿 기입)

    요청 본문(Form Data):
      - template_id: str (예: hometax_official)
      - issue_date: str (예: 2025-10-01 또는 251001 형식 지원)
      - file_name: str (예: 세금계산서_251001.xlsx)
      - file: file (배달대행사 정산서 파일)
      - options: dict (선택)
    동작:
      - 로그인/토큰 체크 후 1토큰 사용
      - 업로드된 파일에서 공급받는자 정보 추출
      - 금액 정보 추출 (요금합계, 부가세)
      - 홈텍스 템플릿에 공급자 + 공급받는자 정보 기입
    """
    user_id, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response
    
    # CSRF 토큰 검증 (간단히 처리)
    csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not csrf_token:
        return error('보안 토큰이 없습니다. 다시 시도해주세요.', status=403)

    # Form Data에서 파라미터 추출
    template_id = (request.form.get('template_id') or 'hometax_official').strip()
    issue_date_raw = (request.form.get('issue_date') or '').strip()
    file_name = (request.form.get('file_name') or '').strip()
    industry_type = (request.form.get('industry_type') or 'delivery').strip()
    guidelines_json = request.form.get('guidelines', '{}')
    
    # 업종별 지침 파싱
    try:
        import json
        guidelines = json.loads(guidelines_json)
        logger.info(f"활성화된 지침: {guidelines.get('name', 'Unknown')}")
    except:
        guidelines = {}
    
    # 파일 업로드 확인
    if 'file' not in request.files:
        return error('배달대행사 정산서 파일을 업로드해주세요', status=400)
    
    uploaded_file = request.files['file']
    if uploaded_file.filename == '':
        return error('파일이 선택되지 않았습니다', status=400)

    if not issue_date_raw:
        return error('전자세금일자를 선택하세요', status=400)
    if not file_name:
        return error('파일명을 입력하세요', status=400)

    # issue_date 정규화: "25년10월01일" 또는 "251001" 또는 ISO 모두 수용 → ISO(YYYY-MM-DD)
    def normalize_issue_date(s: str) -> str:
        try:
            # 251001 형태
            if len(s) == 6 and s.isdigit():
                yy = int(s[0:2])
                mm = int(s[2:4])
                dd = int(s[4:6])
                yyyy = 2000 + yy
                return f"{yyyy:04d}-{mm:02d}-{dd:02d}"
            # 25년10월01일 형태
            if '년' in s and '월' in s and '일' in s:
                yy = int(s.split('년')[0][-2:])
                rest = s.split('년')[1]
                mm = int(rest.split('월')[0])
                dd = int(rest.split('월')[1].split('일')[0])
                yyyy = 2000 + yy
                return f"{yyyy:04d}-{mm:02d}-{dd:02d}"
            # ISO 날짜 시도
            return datetime.fromisoformat(s).date().isoformat()
        except Exception:
            return ''

    issue_date = normalize_issue_date(issue_date_raw)
    if not issue_date:
        return error('전자세금일자 형식이 올바르지 않습니다', status=400)

    # 사용자 정보 로드 → 공급자 정보 자동 매핑
    with get_conn() as conn:
        user = conn.execute(
            """
            SELECT username, email, company_name, business_number,
                   representative_name, phone, address, business_type, business_category,
                   token_balance, COALESCE(tokens_used, 0) AS tokens_used, plan_type
            FROM users WHERE id = ?
            """,
            (user_id,)
        ).fetchone()

        if not user:
            return error('사용자를 찾을 수 없습니다', status=404)
    
    # ============================================
    # 핵심 변경 1: 파일을 먼저 저장
    # ============================================
    # 파일을 임시 디렉토리에 저장
    temp_file_path = save_uploaded_file(uploaded_file)
    
    # ============================================
    # 핵심 변경 2: 템플릿 건수 정밀 계산
    # ============================================
    template_count = calculate_template_count(temp_file_path, industry_type)
    
    if template_count == 0:
        # 임시 파일 정리
        cleanup_temp_file(temp_file_path)
        return error('파일에서 템플릿 건수를 계산할 수 없습니다. 파일 형식을 확인해주세요.', status=400)
    
    logger.info(f"템플릿 건수 계산 완료: {template_count}개")
    
    # ============================================
    # 핵심 변경 2: VIP/GoldVIP 무제한 처리
    # ============================================
    # Gold VIP 회원은 무제한 사용 가능
    logger.info(f"사용자 플랜 확인 시작: user_id={user_id}")
    subscription_info = get_user_subscription(user_id)
    logger.info(f"구독 정보: {subscription_info}")
    
    is_unlimited = is_unlimited_user(user_id)
    logger.info(f"is_unlimited_user 결과: {is_unlimited}")
    
    if is_unlimited:
        required_tokens = 0  # Gold VIP는 토큰 차감 안함
        logger.info(f"GoldVIP 무제한 사용: 템플릿 {template_count}개 변환")
    else:
        # VIP/Premium VIP/Free 회원: 템플릿 건수만큼 토큰 필요
        required_tokens = template_count
        logger.info(f"템플릿 건수: {template_count}개, 필요 토큰: {required_tokens}개")
    
    # ============================================
    # 핵심 변경 3: 토큰 잔량 정밀 확인 (activity_logs 기반)
    # ============================================
    # [버그 수정] users 테이블 대신 activity_logs 기반으로 정확한 토큰 잔량 계산
    # get_token_summary_v2()와 동일한 로직 사용
    with get_conn() as conn:
        conn.row_factory = sqlite3.Row
        summary = conn.execute(
            """
            WITH last_reset AS (
                -- 1. 가장 최근의 TOKEN_RESET_BY_ADMIN 이벤트의 timestamp를 찾는다.
                SELECT MAX(timestamp) as reset_time
                FROM activity_logs
                WHERE user_id = ? AND activity_type = 'TOKEN_RESET_BY_ADMIN'
                  AND COALESCE(is_deleted, 0) = 0  -- [버그 수정] 삭제된 레코드 제외
            )
            SELECT
                -- 2. 해당 리셋 시간 이후의 모든 로그만을 대상으로 집계한다.
                -- 단, TOKEN_RESET_BY_ADMIN의 token_change는 사용량 계산에서 제외한다.
                COALESCE(SUM(CASE WHEN al.token_change > 0 AND al.activity_type != 'TOKEN_RESET_BY_ADMIN' THEN al.token_change ELSE 0 END), 0) as total_charged,
                COALESCE(SUM(CASE WHEN al.token_change < 0 AND al.activity_type != 'TOKEN_RESET_BY_ADMIN' THEN ABS(al.token_change) ELSE 0 END), 0) as total_used
            FROM activity_logs al, last_reset lr
            WHERE al.user_id = ?
              AND (lr.reset_time IS NULL OR al.timestamp >= lr.reset_time)
              AND COALESCE(al.is_deleted, 0) = 0;  -- [버그 수정] 삭제된 레코드 제외
            -- 만약 리셋 기록이 없다면 (lr.reset_time IS NULL), 모든 로그를 포함한다.
            """,
            (user_id, user_id)
        ).fetchone()
        
        total_tokens = summary['total_charged'] if summary else 0
        used_tokens = summary['total_used'] if summary else 0
        available_tokens = total_tokens - used_tokens
    
    if not is_unlimited and available_tokens < required_tokens:
        # 부족한 토큰 정확히 계산
        shortage = required_tokens - available_tokens
        missing_tokens = shortage
        
        logger.warning(
            f"토큰 부족 - 템플릿 {template_count}개, 필요 {required_tokens}토큰, "
            f"보유 {available_tokens}토큰, 부족 {shortage}토큰"
        )
        
        return error(
            f'토큰이 부족합니다. 템플릿 {template_count}개 생성에 {required_tokens}토큰이 필요하지만, '
            f'현재 {available_tokens}토큰을 보유하고 있어 {shortage}토큰이 부족합니다. '
            f'필요한 토큰({shortage}개 × 200원 = {shortage * 200}원)을 구매한 후 다시 시도해주세요.',
            status=400,
            data={
                'template_count': template_count,
                'required_tokens': required_tokens,
                'available_tokens': available_tokens,
                'shortage': shortage,
                'estimated_cost': shortage * 200  # 1토큰 = 200원
            }
        )
    
    # ============================================
    # 핵심 변경 4: 토큰은 변환 후 실제 생성된 수만큼만 차감
    # ============================================
    # 변환 전에는 차감하지 않고 잔량만 확인함
    logger.info(
        f"토큰 잔량 확인 완료: 필요 {required_tokens}토큰, 보유 {available_tokens}토큰"
    )
    
    # 변환 시작 시간 기록
    import time
    conversion_start_time = time.time()
    logger.info(f"변환 시작: {time.strftime('%Y-%m-%d %H:%M:%S')} - 사용자 {user_id}")

    # ============================================
    # 골드 회원 공급자 선택 분기
    # ============================================
    # selectedCustomerId 파라미터 확인
    selected_customer_id = request.form.get('selectedCustomerId', '').strip()
    
    supplier = None
    
    # 골드 회원이고 고객을 선택한 경우
    user_plan_type = user['plan_type'] if 'plan_type' in user.keys() else None
    if selected_customer_id and user_plan_type in ['gold', 'gold-vip']:
        logger.info(f"골드 고객 선택됨: customer_id={selected_customer_id}")
        
        with get_conn() as conn:
            customer = conn.execute(
                "SELECT * FROM gold_customers WHERE id = ? AND user_id = ? AND is_deleted = 0",
                (int(selected_customer_id), user_id)
            ).fetchone()
            
            if customer:
                import json
                business_kind = customer['business_kind'] if 'business_kind' in customer.keys() else '{}'
                try:
                    business_kind_dict = json.loads(business_kind) if isinstance(business_kind, str) else business_kind
                except:
                    business_kind_dict = {}
                
                supplier = {
                    'supplier_name': customer['company_name'] if 'company_name' in customer.keys() else '',
                    'supplier_business_number': customer['business_number'] if 'business_number' in customer.keys() else '',
                    'supplier_representative': customer['representative_name'] if 'representative_name' in customer.keys() else '',
                    'supplier_email': customer['email'] if 'email' in customer.keys() else '',
                    'supplier_address': customer['address'] if 'address' in customer.keys() else '',
                    'supplier_business_type': business_kind_dict.get('업태', ''),
                    'supplier_business_category': business_kind_dict.get('종목', ''),
                }
                logger.info(f"골드 고객 정보 적용: {supplier['supplier_name']}")
    
    # 골드 고객 미선택 또는 비골드 회원: 기본 프로필 공급자 사용
    if not supplier:
        supplier = {
            'supplier_name': user['company_name'] or user['username'],
            'supplier_business_number': user['business_number'] or '',
            'supplier_representative': user['representative_name'] or '',
            'supplier_email': user['email'] or '',
            'supplier_address': user['address'] or '',
            'supplier_business_type': user['business_type'] if 'business_type' in user.keys() else '',
            'supplier_business_category': user['business_category'] if 'business_category' in user.keys() else '',
        }
        logger.info(f"기본 프로필 공급자 사용: {supplier['supplier_name']}")
    
    # 사용자 정보를 절대지침 시스템에 전달
    user_info = {
        'user_id': user_id,  # 사용자 ID 추가
        'company_name': user['company_name'] or user['username'],
        'business_number': user['business_number'] or '',
        'representative': user['representative_name'] or '',
        'email': user['email'] or '',
        'address': user['address'] or '',
        'phone': user['phone'] if 'phone' in user.keys() else '',
        'business_type': user['business_type'] or '',
        'business_category': user['business_category'] or '',
    }

    # 파일은 이미 저장됨 (temp_file_path에 저장되어 있음)

    # 변환 엔진 실행
    try:
        # 요청별 새로운 변환 엔진 인스턴스 생성 (상태 격리)
        conversion_engine = ConversionEngine()
        
        # 전체 변환 프로세스 실행 (상태 격리된 인스턴스 사용)
        conversion_result = conversion_engine.convert_file(
            uploaded_file_path=temp_file_path,
            supplier_info=supplier,
            template_id=template_id,
            industry_type=industry_type,
            guidelines=guidelines,
            issue_date=issue_date,
            file_name=file_name,
            user_info=user_info
        )
        
        if not conversion_result['success']:
            # 임시 파일 정리
            cleanup_temp_file(temp_file_path)
            return error(f"변환 실패: {conversion_result.get('error_message', '알 수 없는 오류')}", status=500)
        
        # 임시 파일 정리
        cleanup_temp_file(temp_file_path)
        
        # ============================================
        # 핵심 변경: 기록관을 통한 활동 로그 기록 및 토큰 업데이트
        # ============================================
        # DB 트랜잭션 내에서 활동 로그 기록 및 토큰 업데이트
        try:
            with get_conn() as conn:
                cursor = conn.cursor()
                
                # 사용자 정보 재조회 (최신 토큰 잔액 확인)
                user_current = cursor.execute(
                    """
                    SELECT id, plan_type, token_balance, COALESCE(tokens_used, 0) AS tokens_used
                    FROM users WHERE id = ?
                    """,
                    (user_id,)
                ).fetchone()
                
                if not user_current:
                    logger.error(f"사용자 정보를 찾을 수 없습니다: user_id={user_id}")
                    return error('사용자 정보를 찾을 수 없습니다', status=404)
                
                # --- [수정] '기록관' 호출 방식을 새로운 범용 함수에 맞게 변경 ---
                
                # 1. 정보 추출
                user_id_for_activity = user_current['id']
                plan_type = user_current['plan_type'] or 'free'
                token_balance_before = user_current['token_balance'] or 0
                total_recipients = conversion_result.get('total_recipients', 0)
                
                # 2. '경제 헌법' 적용
                potential_cost = total_recipients * -1
                token_change = 0
                if plan_type not in ['unlimited', 'gold', 'gold-vip']:
                    token_change = potential_cost
                token_balance_after = token_balance_before + token_change
                
                # 3. 범용 activity_data 생성
                activity_data = {
                    'user_id': user_id_for_activity,
                    'performed_by_id': user_id_for_activity,
                    'performed_by_type': 'USER',
                    'activity_type': 'FILE_CONVERT',
                    'details': {
                        "filename": file_name,
                        "extracted_rows": total_recipients,
                        "cost_policy": "1_token_per_row(temp)"
                    },
                    'token_change': token_change,
                    'potential_cost': potential_cost,
                    'token_balance_before': token_balance_before,
                    'token_balance_after': token_balance_after,
                    'user_plan_snapshot': plan_type
                }
                
                # 4. 새로운 범용 '기록관' 호출
                record_activity(cursor, activity_data)
                
                # 트랜잭션 커밋
                conn.commit()
                logger.info(f"활동 로그 기록 완료: user_id={user_id}, file_name={file_name}")
                
        except Exception as activity_error:
            logger.error(f"활동 로그 기록 중 오류 발생: {str(activity_error)}")
            # 활동 로그 기록 실패는 치명적이지 않으므로 계속 진행
            # 하지만 경고 로그는 남김
            import traceback
            traceback.print_exc()
        
        # ============================================
        # 핵심 변경: 연동 모듈을 통한 토큰 차감
        # ============================================
        token_processor = TokenDeductionProcessor()
        token_result = token_processor.process_token_deduction(
            user_id=user_id,
            is_unlimited=is_unlimited,
            conversion_result=conversion_result
        )

        if not token_result.get('success'):
            logger.error(f"토큰 차감 실패: {token_result.get('message')}")
            return error(token_result.get('message', '토큰 처리 중 오류가 발생했습니다'), status=500)

        logger.info(f"토큰 차감 결과: {token_result['message']}")
        
        # 변환 완료 시간 기록 및 실행 시간 계산
        conversion_end_time = time.time()
        execution_time = round(conversion_end_time - conversion_start_time, 2)
        logger.info(f"변환 완료: {time.strftime('%Y-%m-%d %H:%M:%S')} - 실행시간: {execution_time}초 - 사용자 {user_id}")
        
        # 세션에 변환 결과 저장
        session['last_conversion_result'] = conversion_result
        session['last_file_name'] = file_name  # 다운로드 파일명 저장
        
        tokens_payload = {
            'total_granted': token_result.get('total_granted', user['token_balance'] or 0),
            'total_used': token_result.get('tokens_used_after', user['tokens_used'] or 0),
            'available_tokens': token_result.get('available_tokens_after', (user['token_balance'] or 0) - (user['tokens_used'] or 0)),
            'templates_created': token_result.get('recipient_count', 0)
        }

        return success('변환 완료', data={
            'conversion_result': conversion_result,
            'download_url': url_for('conversion.download_converted', _external=False),
            'download_filename': file_name,
            'detailed_stats': conversion_result.get('detailed_stats', {}),
            'tokens': tokens_payload
        })
        
    except Exception as e:
        # 변환 전에 토큰을 차감하지 않았으므로 환불 불필요
        logger.error(f"변환 실패: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # 임시 파일 정리
        if 'temp_file_path' in locals():
            cleanup_temp_file(temp_file_path)
        return error(f'변환 처리 중 오류 발생: {str(e)}', status=500)


@conversion_bp.route('/api/convert/download', methods=['GET'])
def download_converted():
    """변환된 홈텍스 파일 다운로드

    - 변환 결과 파일이 1개인 경우: XLSX 파일을 직접 다운로드로 반환
    - 변환 결과 파일이 2개 이상인 경우: ZIP으로 묶어 반환 (기존 동작 유지)
    """
    _, guard_response = ensure_login_for_json()
    if guard_response is not None:
        return guard_response

    conversion_result = session.get('last_conversion_result')
    if not conversion_result or not conversion_result.get('success'):
        return error('다운로드할 변환 결과가 없습니다', status=404)

    try:
        from flask import send_file

        files = [p for p in conversion_result.get('files', []) if os.path.exists(p)]

        # 사용자가 다운로드 모드를 선택할 수 있도록 쿼리 파라미터 지원
        # mode=auto(기본): 1개면 단일, 2개 이상이면 ZIP
        # mode=zip: 개수와 무관하게 ZIP으로 묶어서 제공
        mode = (request.args.get('mode') or 'auto').lower()

        # 파일이 하나라면 ZIP이 아닌 단일 파일로 내려준다 (또는 zip 강제 시 ZIP)
        if len(files) == 1 and mode != 'zip':
            single_path = files[0]
            filename = os.path.basename(single_path)
            # xlsx/xlsm 등 확장자에 따라 MIME 설정 (기본 xlsx)
            ext = os.path.splitext(filename)[1].lower()
            if ext in ['.xlsx', '.xltx']:
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif ext in ['.xlsm']:
                mimetype = 'application/vnd.ms-excel.sheet.macroEnabled.12'
            else:
                mimetype = 'application/octet-stream'

            return send_file(
                single_path,
                as_attachment=True,
                download_name=filename,
                mimetype=mimetype
            )

        # 2개 이상 또는 zip 강제 시 ZIP으로 제공
        import zipfile
        import io

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_path in files:
                filename = os.path.basename(file_path)
                zip_file.write(file_path, filename)
        zip_buffer.seek(0)

        # 사용자 입력 파일명 사용 (쿼리 파라미터 우선, 세션 차선, 없으면 기본값)
        user_file_name = (request.args.get('filename') or session.get('last_file_name', '')).strip()
        zip_filename = '홈텍스_일괄등록_파일들.zip'  # 기본값
        
        if user_file_name:
            # 확장자 제거 후 .zip 추가
            base_name = user_file_name.rsplit('.', 1)[0] if '.' in user_file_name else user_file_name
            # Windows 금지 문자 제거
            import re
            base_name = re.sub(r'[\\/:*?"<>|]', '', base_name).strip()
            if base_name:  # 제거 후 빈 문자열이 아니면
                zip_filename = f"{base_name}.zip"
        
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
    except Exception as e:
        return error(f'다운로드 처리 중 오류 발생: {str(e)}', status=500)

    # 1) XLSX 시도 (openpyxl)
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
        import io

        # 내장 템플릿 로드 시도
        template_id = (payload.get('template_id') or 'hometax_official')
        template_info = template_manager.get_template_info(template_id)
        wb = None
        ws = None
        if template_info:
            tpath = template_manager.get_template_path(template_id)
            if tpath:
                try:
                    wb = load_workbook(tpath)
                    sheet_name = template_info.get('sheet_name') or wb.active.title
                    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                except Exception:
                    wb = None
                    ws = None

        # 템플릿 로드 실패 시 스켈레톤 워크북 생성
        if wb is None or ws is None:
            wb = Workbook()
            ws = wb.active
            ws.title = '공급자정보'

        # 핵심지침 적용: 헤더 스캔(6행) → 정규화/별칭 매핑 → 열 매핑 → 7행 기입
        header_row_index = 6
        target_row_index = 7

        # 1) 헤더 정규화 함수: 괄호/따옴표/공백/기호 제거, 소문자화
        import re
        def normalize_header(text: str) -> str:
            if text is None:
                return ''
            s = str(text)
            # 괄호 안 설명 제거 ((), [] 안의 내용 포함)
            s = re.sub(r"\(.*?\)", "", s)
            s = re.sub(r"\[.*?\]", "", s)
            # 따옴표 제거
            s = s.replace('"', '').replace("'", '')
            # 특수기호 제거
            s = re.sub(r"[^0-9A-Za-z가-힣]+", "", s)
            # 공백/대소문자 정규화
            s = s.strip().lower()
            return s

        # 2) 별칭 사전: 정규화된 헤더명을 표준 키로 매핑
        # 표준 키(사전에 사용할 키)는 원문을 유지
        canonical_headers = [
            '작성일자',
            '공급자 상호',
            '공급자 성명',
            '공급자 등록번호',
            '공급자 종사업장번호',
            '공급자 사업장주소',
            '공급자 업태',
            '공급자 종목',
            '공급자 이메일',
        ]

        alias_map = {
            normalize_header('작성일자'): '작성일자',
            normalize_header('공급자상호'): '공급자 상호',
            normalize_header('공급자 성명'): '공급자 성명',
            normalize_header('공급자등록번호'): '공급자 등록번호',
            normalize_header('공급자 등록번호'): '공급자 등록번호',
            normalize_header('공급자 종사업장번호'): '공급자 종사업장번호',
            normalize_header('공급자사업장주소'): '공급자 사업장주소',
            normalize_header('공급자 사업장주소'): '공급자 사업장주소',
            normalize_header('공급자 업태'): '공급자 업태',
            normalize_header('공급자종목'): '공급자 종목',
            normalize_header('공급자 이메일'): '공급자 이메일',
        }

        # 3) 시트 헤더 스캔 → 표준 키로 열 인덱스 매핑
        headers_by_canonical: dict[str, int] = {}
        try:
            max_col = ws.max_column or 100
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=header_row_index, column=col).value
                if cell_value is None:
                    continue
                norm = normalize_header(cell_value)
                if not norm:
                    continue
                if norm in alias_map:
                    canonical = alias_map[norm]
                    headers_by_canonical[canonical] = col
        except Exception:
            headers_by_canonical = {}

        # 4) 필수 헤더 존재 확인(정규화/별칭 적용 후)
        missing = [h for h in canonical_headers if h not in headers_by_canonical]
        if missing:
            return error(f"템플릿 헤더 누락: {', '.join(missing)}", status=422)

        # 날짜 형식: YYYY-MM-DD → 251001(YYMMDD)
        issue_iso = payload.get('issue_date') or ''
        try:
            yy = issue_iso[2:4]
            mm = issue_iso[5:7]
            dd = issue_iso[8:10]
            issue_compact = f"{yy}{mm}{dd}"
        except Exception:
            issue_compact = ''

        # 작성일자(B6 헤더의 열)에 7행부터 기입(현재 1건이므로 한 셀)
        ws.cell(row=target_row_index, column=headers_by_canonical['작성일자'], value=issue_compact)

        # 공급자 정보 기입(7행)
        supplier = payload.get('supplier', {})
        def write_if(header_name: str, value: str):
            try:
                ws.cell(row=target_row_index, column=headers_by_canonical[header_name], value=value)
            except Exception:
                pass

        write_if('공급자 상호', supplier.get('supplier_name', ''))
        write_if('공급자 성명', supplier.get('supplier_representative', ''))
        write_if('공급자 등록번호', supplier.get('supplier_business_number', ''))
        write_if('공급자 종사업장번호', supplier.get('supplier_branch_number', ''))
        write_if('공급자 사업장주소', supplier.get('supplier_address', ''))
        write_if('공급자 업태', supplier.get('supplier_business_type', ''))
        write_if('공급자 종목', supplier.get('supplier_business_item', ''))
        write_if('공급자 이메일', supplier.get('supplier_email', ''))

        # 절대값 규칙: A7=01, W7=30, BG7=01
        try:
            ws['A7'] = '01'
            ws['W7'] = '30'
            ws['BG7'] = '01'
        except Exception:
            pass

        # 템플릿이 없었던 경우에만 참고용 표도 작성
        mapping = payload.get('mapping', {})
        if ws.title == '공급자정보' and not mapping:
            ws.cell(row=1, column=1, value='field')
            ws.cell(row=1, column=2, value='value')
            supplier = payload.get('supplier', {})
            order = ['supplier_name','supplier_business_number','supplier_representative','supplier_phone','supplier_email','supplier_address']
            r = 2
            for k in order:
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=supplier.get(k, ''))
                r += 1
            # 간단한 너비 조정
            for col in range(1, 3):
                length = 12
                for row in ws.iter_rows(min_row=1, max_row=r-1, min_col=col, max_col=col):
                    for c in row:
                        if c.value:
                            length = max(length, len(str(c.value)) + 2)
                ws.column_dimensions[get_column_letter(col)].width = length

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        fname = (payload.get('file_name') or 'converted.xlsx')
        if not fname.lower().endswith(('.xlsx', '.xls')):
            base = fname.rsplit('.', 1)[0] if '.' in fname else fname
            fname = base + '.xlsx'

        from flask import Response
        # RFC 5987: ASCII-only fallback + UTF-8 encoded filename*
        ascii_fallback = 'download.xlsx'
        encoded = quote(fname)
        return Response(
            bio.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f"attachment; filename={ascii_fallback}; filename*=UTF-8''{encoded}",
                'Cache-Control': 'no-store'
            }
        )
    except Exception:
        # 2) 폴백: CSV 제공
        import io, csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['field', 'value'])

        supplier = payload.get('supplier', {})
        for k in ['supplier_name','supplier_business_number','supplier_representative','supplier_phone','supplier_email','supplier_address']:
            writer.writerow([k, supplier.get(k, '')])

        writer.writerow([])
        writer.writerow(['mapping_cell', 'mapped_value'])
        mapping = payload.get('mapping', {})
        for cell, val in mapping.items():
            writer.writerow([cell, val])

        csv_text = '\ufeff' + buf.getvalue()

        fname = (payload.get('file_name') or 'converted.csv')
        if fname.lower().endswith(('.xlsx', '.xls')):
            fname = fname.rsplit('.', 1)[0] + '.csv'
        elif not fname.lower().endswith('.csv'):
            fname = fname + '.csv'

        from flask import Response
        # RFC 5987: ASCII-only fallback + UTF-8 encoded filename*
        ascii_fallback = 'download.csv'
        encoded = quote(fname)
        return Response(
            csv_text,
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f"attachment; filename={ascii_fallback}; filename*=UTF-8''{encoded}",
                'Cache-Control': 'no-store'
            }
        )

@conversion_bp.route('/api/security/validation/test', methods=['POST'])
def test_file_validation():
    """파일 검증 테스트 API"""
    _, guard_response = ensure_admin_for_json()
    if guard_response is not None:
        return guard_response
    
    try:
        data = request.get_json(silent=True) or {}
        file_path = data.get('file_path')
        target_folder = data.get('target_folder')
        
        if not file_path or not target_folder:
            return error('file_path와 target_folder가 필요합니다', status=400)
        
        # 경로 검증 테스트
        is_valid, message = file_validator.validate_destination_path(file_path, target_folder)
        
        return success('파일 검증 테스트 완료', data={
            'file_path': file_path,
            'target_folder': target_folder,
            'is_valid': is_valid,
            'message': message
        })
        
    except Exception as e:
        return error(f'파일 검증 테스트 실패: {str(e)}', status=500)

