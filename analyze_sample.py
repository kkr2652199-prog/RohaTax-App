#!/usr/bin/env python3
"""
sample_invoice4.xlsx 파일 분석 스크립트
"""

import openpyxl
import pandas as pd

def analyze_sample_file():
    """sample_invoice4.xlsx 파일 분석"""
    file_path = 'tests/input/sample_invoice4.xlsx'
    
    try:
        # Excel 파일 로드
        wb = openpyxl.load_workbook(file_path)
        print(f"📄 파일: {file_path}")
        print(f"📋 시트 목록: {wb.sheetnames}")
        print("=" * 80)
        
        # 각 시트 분석
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n🔍 시트: '{sheet_name}'")
            print(f"   행 수: {sheet.max_row}, 열 수: {sheet.max_column}")
            
            # 첫 5행 데이터 표시
            print("   첫 5행 데이터:")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(11, sheet.max_column + 1)):  # 최대 10열까지만
                    cell_value = sheet.cell(row, col).value
                    row_data.append(str(cell_value)[:20] if cell_value else '')  # 20자로 제한
                print(f"     행 {row}: {row_data}")
            
            # 컬럼명 추출 (첫 번째 행)
            if sheet.max_row > 0:
                headers = []
                for col in range(1, min(21, sheet.max_column + 1)):  # 최대 20열까지만
                    header = sheet.cell(1, col).value
                    headers.append(str(header) if header else f'열{col}')
                print(f"   헤더: {headers}")
        
        print("\n" + "=" * 80)
        
        # pandas로도 읽어보기
        print("\n📊 pandas로 읽은 데이터:")
        for sheet_name in wb.sheetnames:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                print(f"\n시트 '{sheet_name}':")
                print(f"  컬럼: {list(df.columns)}")
                print(f"  데이터 형태: {df.shape}")
                print(f"  첫 3행:")
                print(df.head(3).to_string())
            except Exception as e:
                print(f"  pandas 읽기 실패: {e}")
                
    except Exception as e:
        print(f"❌ 파일 분석 실패: {e}")

if __name__ == "__main__":
    analyze_sample_file()
