"""헤더 탐지 로직 전담 모듈."""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from .scoring_utils import (
    calculate_data_density,
    calculate_csv_data_density,
    count_csv_matched_fields,
)

logger = logging.getLogger(__name__)


def detect_header_row(sheet: Worksheet) -> int:
    """엑셀 시트에서 헤더 행을 감지한다."""

    max_row = sheet.max_row
    max_col = sheet.max_column

    title_rows = detect_title_rows(sheet, max_col)
    logger.info(f"🔍 제목/부제목 행 감지: {title_rows}")

    header_candidates: List[Tuple[int, float]] = []

    for row in range(1, min(1001, max_row + 1)):
        if row in title_rows:
            continue

        data_density = calculate_data_density(sheet, row, max_col)
        header_candidates.append((row, data_density))

    if not header_candidates:
        logger.warning("헤더 후보가 없습니다. 기본 헤더 행 사용")
        return 1

    header_candidates.sort(key=lambda x: (x[1], x[0]), reverse=True)
    best_header_row = header_candidates[0][0]

    logger.info(f"헤더 후보 분석: {header_candidates[:5]}...")
    logger.info(f"선택된 헤더 행: {best_header_row}")

    return best_header_row


def detect_csv_header_row(
    df: pd.DataFrame,
    required_keywords: Dict[str, List[str]],
) -> int:
    """CSV 파일에서 헤더 행을 감지한다."""

    if len(df) == 0:
        return 0

    header_candidates: List[Dict[str, Any]] = []
    scan_rows = min(10, len(df))

    for row_idx in range(scan_rows):
        data_density = calculate_csv_data_density(df, row_idx)
        matched_fields = count_csv_matched_fields(
            df,
            row_idx,
            required_keywords,
        )

        field_match_score = (matched_fields / 5) * 0.8
        density_score = data_density * 0.2
        header_score = field_match_score + density_score

        if matched_fields >= 3:
            header_candidates.append(
                {
                    'row': row_idx,
                    'data_density': data_density,
                    'matched_fields': matched_fields,
                    'header_score': header_score,
                }
            )
            logger.debug(
                "지능앱 CSV 헤더 후보: 행 %d, 점수 %.3f, 매칭 %d개",
                row_idx,
                header_score,
                matched_fields,
            )

    if not header_candidates:
        density_scores = [
            (row, calculate_csv_data_density(df, row)) for row in range(scan_rows)
        ]
        density_scores.sort(key=lambda x: (x[1], -x[0]), reverse=True)
        best_header_row = density_scores[0][0]
        logger.warning(
            "지능앱 CSV 헤더 감지: 매칭된 헤더가 없어 데이터 밀도로 선택 - 행 %d",
            best_header_row,
        )
        return best_header_row

    header_candidates.sort(
        key=lambda x: (x['matched_fields'], x['header_score'], x['row']),
        reverse=True,
    )
    best_header = header_candidates[0]
    best_header_row = best_header['row']
    logger.info(
        "지능앱 CSV 헤더 감지: 최적 헤더 선택 - 행 %d (점수: %.3f, 매칭: %d개)",
        best_header_row,
        best_header['header_score'],
        best_header['matched_fields'],
    )

    return best_header_row


def detect_title_rows(sheet: Worksheet, max_col: int) -> List[int]:
    title_rows: List[int] = []

    for row in range(1, min(21, sheet.max_row + 1)):
        text_pattern = analyze_text_pattern(sheet, row, max_col)
        if is_title_pattern(text_pattern):
            title_rows.append(row)
            logger.debug(
                "제목/부제목 행 감지: %d행 - %s",
                row,
                text_pattern,
            )

    return title_rows


def analyze_text_pattern(
    sheet: Worksheet,
    row: int,
    max_col: int,
) -> Dict[str, int]:
    text_count = 0
    number_count = 0
    empty_count = 0
    long_text_count = 0

    max_col = min(max_col, 50)

    for col in range(1, max_col + 1):
        cell_value = sheet.cell(row=row, column=col).value

        if cell_value is None or str(cell_value).strip() == '':
            empty_count += 1
        elif isinstance(cell_value, (int, float)):
            number_count += 1
        else:
            text_count += 1
            if len(str(cell_value).strip()) > 10:
                long_text_count += 1

    return {
        'text_count': text_count,
        'number_count': number_count,
        'empty_count': empty_count,
        'long_text_count': long_text_count,
    }


def is_title_pattern(pattern: Dict[str, int]) -> bool:
    return (
        pattern['text_count'] >= 3
        and pattern['long_text_count'] >= 1
        and pattern['number_count'] <= 1
    )


def find_actual_data_range(sheet: Worksheet) -> Tuple[int, int]:
    actual_max_row = 1
    actual_max_col = 1

    try:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    if cell.row > actual_max_row:
                        actual_max_row = cell.row
                    if cell.column > actual_max_col:
                        actual_max_col = cell.column

        actual_max_row = max(actual_max_row, 2)
        actual_max_col = max(actual_max_col, 5)
        logger.info(
            "실제 데이터 범위 감지: 행 %d, 열 %d",
            actual_max_row,
            actual_max_col,
        )
    except Exception as exc:
        logger.warning("실제 데이터 범위 감지 중 오류: %s", exc)
        actual_max_row = min(sheet.max_row, 1000)
        actual_max_col = min(sheet.max_column, 50)

    return actual_max_row, actual_max_col


def get_actual_data_range(sheet: Worksheet) -> Tuple[int, int]:
    """공개용: 실제 데이터 범위를 반환한다."""
    return find_actual_data_range(sheet)
