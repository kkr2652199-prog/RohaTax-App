"""
시트 평가 및 선택 시스템 모듈
Excel 워크북의 모든 시트를 평가하고 최적의 시트를 선택하는 시스템
"""

import openpyxl
from typing import Dict, List, Any, Optional, Tuple
import logging
from .header_detector import HeaderDetector
from ..industry_config_loader import industry_config_loader

logger = logging.getLogger(__name__)


class SheetEvaluator:
    """지능형 시트 평가 및 선택 시스템"""
    
    def __init__(self):
        self.logger = logger
        self.header_detector = HeaderDetector()
    
    def inspect_all_sheets(self, workbook) -> Optional[Dict[str, Any]]:
        """
        지능앱 핵심 기술: 모든 시트를 검열하여 최적의 시트 선택
        배달대행사 정산서의 5가지 필수 컬럼을 기준으로 시트 평가
        
        Args:
            workbook: openpyxl Workbook 객체
            
        Returns:
            Dict: 최적의 시트 정보 또는 None
        """
        sheet_results = []
        best_result = None
        best_score = -1

        # Load delivery scoring config (안전한 로딩)
        try:
            delivery_cfg = industry_config_loader.get_industry_config('delivery') or {}
            rules = (delivery_cfg.get('extraction_rules') or {})
            scoring = (rules.get('sheet_scoring') or {})
            weights = (scoring.get('weights') or {
                'business_number': 30,
                'representative': 10,
                'address': 30,
                'email': 20,
                'store_name': 10,
            })
            thresholds = (scoring.get('thresholds') or {'pass': 80, 'candidate': 70})
            override_all5 = bool(scoring.get('override_select_if_all_5_found', True))
        except Exception as e:
            self.logger.warning(f"업종 설정 로딩 실패, 기본값 사용: {str(e)}")
            weights = {
                'business_number': 30,
                'representative': 10,
                'address': 30,
                'email': 20,
                'store_name': 10,
            }
            thresholds = {'pass': 80, 'candidate': 70}
            override_all5 = True

        # 모든 시트를 결정적 순서로 처리
        sorted_sheet_names = sorted(workbook.sheetnames)
        
        for sheet_name in sorted_sheet_names:
            try:
                sheet = workbook[sheet_name]
                self.logger.info(f"📊 시트 '{sheet_name}' 평가 시작")
                
                # 시트 평가 실행
                sheet_result = self.evaluate_sheet(sheet, sheet_name, weights, thresholds)
                
                if sheet_result:
                    sheet_results.append(sheet_result)
                    self.logger.info(f"✅ 시트 '{sheet_name}' 평가 완료: 점수 {sheet_result.get('score', 0):.2f}")
                    
                    # 최고 점수 시트 업데이트
                    if sheet_result.get('score', 0) > best_score:
                        best_score = sheet_result.get('score', 0)
                        best_result = sheet_result
                else:
                    self.logger.info(f"❌ 시트 '{sheet_name}' 평가 실패")
                    
            except Exception as e:
                self.logger.error(f"시트 '{sheet_name}' 평가 중 오류: {str(e)}")
                continue

        # 결과 요약 로그
        if sheet_results:
            self.logger.info(f"📈 총 {len(sheet_results)}개 시트 평가 완료")
            for result in sorted(sheet_results, key=lambda x: x.get('score', 0), reverse=True)[:3]:
                self.logger.info(f"  - {result.get('sheet_name', 'unknown')}: {result.get('score', 0):.2f}점")
        else:
            self.logger.warning("평가된 시트가 없습니다")

        return best_result
    
    def evaluate_sheet(self, sheet, sheet_name: str, weights: Dict, thresholds: Dict) -> Optional[Dict[str, Any]]:
        """
        개별 시트 평가 - 지능앱 기술 적용
        5가지 필수 컬럼 매칭과 데이터 품질을 종합적으로 평가
        
        Args:
            sheet: openpyxl Worksheet 객체
            sheet_name: 시트 이름
            weights: 가중치 딕셔너리
            thresholds: 임계값 딕셔너리
            
        Returns:
            Dict: 시트 평가 결과 또는 None
        """
        try:
            # 🎯 실제 데이터 범위 감지: 각 시트의 실제 마지막 행/열 찾기
            actual_max_row, actual_max_col = self._find_actual_data_range(sheet)
            
            # 지침 반영: 행은 실제 마지막 행까지 사용, 열은 안전 상한 유지(50)
            max_row = actual_max_row
            max_col = min(actual_max_col, 50)
            
            if max_row < 2 or max_col < 5:
                return None
            
            # 지능앱 기술: 헤더 행 후보 검색 (상단 30행까지 확장)
            header_candidates = []
            
            # 지능앱 기술: 스캔 범위 확장 (복잡한 파일 대응)
            scan_rows = min(1000, max_row)  # 최대 1000행까지 스캔 (장문 제목/설명 대응)
            
            for row in range(1, scan_rows + 1):
                try:
                    # 금지어 맵이 제공된 경우 사용
                    matched_fields = self._count_matched_fields(sheet, row, max_col, weights)
                    
                    if matched_fields > 0:
                        # 데이터 품질 점수 계산
                        data_quality = self._calculate_data_quality_score(sheet, row, max_col)
                        
                        # 종합 점수 계산 (매칭 필드 수 + 데이터 품질)
                        total_score = matched_fields + data_quality
                        
                        header_candidates.append({
                            'row': row,
                            'matched_fields': matched_fields,
                            'data_quality': data_quality,
                            'total_score': total_score
                        })
                        
                except Exception as e:
                    self.logger.debug(f"행 {row} 평가 중 오류: {str(e)}")
                    continue
            
            if not header_candidates:
                return None
            
            # 최고 점수 헤더 선택 (결정적 정렬)
            header_candidates.sort(key=lambda x: (x['total_score'], x['matched_fields'], x['row']), reverse=True)
            best_header = header_candidates[0]
            
            # 시트 점수 계산
            sheet_score = self._calculate_sheet_score(sheet, best_header, weights, thresholds)
            
            if sheet_score < thresholds.get('candidate', 70):
                return None
            
            # 가족 데이터 추출
            families = self._extract_families_from_sheet(sheet, best_header['row'], max_col)
            
            return {
                'sheet_name': sheet_name,
                'header_row': best_header['row'],
                'score': sheet_score,
                'matched_fields': best_header['matched_fields'],
                'data_quality': best_header['data_quality'],
                'families': families,
                'family_count': len(families)
            }
            
        except Exception as e:
            self.logger.error(f"시트 '{sheet_name}' 평가 중 오류: {str(e)}")
            return None
    
    def _find_actual_data_range(self, sheet) -> tuple[int, int]:
        """실제 데이터 범위 감지"""
        actual_max_row = 1
        actual_max_col = 1
        
        try:
            # 모든 셀을 순회하여 실제 데이터가 있는 범위 찾기
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        # 실제 데이터가 있는 셀 발견
                        if cell.row > actual_max_row:
                            actual_max_row = cell.row
                        if cell.column > actual_max_col:
                            actual_max_col = cell.column
            
            # 최소값 보장 (헤더가 있을 수 있으므로)
            actual_max_row = max(actual_max_row, 2)
            actual_max_col = max(actual_max_col, 5)
            
        except Exception as e:
            self.logger.warning(f"실제 데이터 범위 감지 중 오류: {str(e)}")
            # 오류 시 기본값 사용
            actual_max_row = min(sheet.max_row, 1000)
            actual_max_col = min(sheet.max_column, 50)
        
        return actual_max_row, actual_max_col
    
    def _count_matched_fields(self, sheet, row: int, max_col: int, weights: Dict) -> int:
        """매칭된 필드 수 계산"""
        try:
            matched_count = 0
            
            # 5형제 키워드 정의
            required_keywords = {
                'business_number': ['사업자등록번호', '사업자번호', '등록번호'],
                'store_name': ['상호명', '상호', '업체명', '가맹점명'],
                'representative': ['대표자', '대표자명'],
                'address': ['주소', '사업장주소'],
                'email': ['이메일', 'email']
            }
            
            # 해당 행의 모든 셀 검사
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    
                    # 각 필드별 키워드 매칭 확인
                    for field_name, keywords in required_keywords.items():
                        if any(keyword.lower() in cell_value for keyword in keywords):
                            matched_count += 1
                            break  # 한 필드는 한 번만 카운트
            
            return matched_count
            
        except Exception as e:
            self.logger.warning(f"매칭 필드 계산 중 오류: {str(e)}")
            return 0
    
    def _calculate_data_quality_score(self, sheet, row: int, max_col: int) -> float:
        """데이터 품질 점수 계산"""
        try:
            total_cells = 0
            filled_cells = 0
            
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                total_cells += 1
                
                if cell.value is not None and str(cell.value).strip():
                    filled_cells += 1
            
            return filled_cells / max(total_cells, 1)
            
        except Exception as e:
            self.logger.warning(f"데이터 품질 점수 계산 중 오류: {str(e)}")
            return 0.0
    
    def _calculate_sheet_score(self, sheet, header_info: Dict, weights: Dict, thresholds: Dict) -> float:
        """시트 종합 점수 계산"""
        try:
            # 기본 점수 (매칭된 필드 수 기반)
            base_score = header_info['matched_fields'] * 20  # 필드당 20점
            
            # 데이터 품질 보너스
            quality_bonus = header_info['data_quality'] * 10
            
            # 총 점수
            total_score = base_score + quality_bonus
            
            return min(total_score, 100.0)  # 최대 100점
            
        except Exception as e:
            self.logger.warning(f"시트 점수 계산 중 오류: {str(e)}")
            return 0.0
    
    def _extract_families_from_sheet(self, sheet, header_row: int, max_col: int) -> List[Dict]:
        """시트에서 가족 데이터 추출"""
        families = []
        
        try:
            # 헤더 행에서 컬럼 매핑 생성
            column_mapping = self._create_column_mapping(sheet, header_row, max_col)
            
            # 데이터 행들 처리
            for row_num in range(header_row + 1, min(sheet.max_row + 1, header_row + 1000)):
                family_data = self._extract_family_from_row(sheet, row_num, column_mapping, max_col)
                if family_data:
                    families.append(family_data)
            
            # 가족 통합 처리
            families = self._merge_family_data(families)
            
        except Exception as e:
            self.logger.warning(f"가족 데이터 추출 중 오류: {str(e)}")
        
        return families
    
    def _create_column_mapping(self, sheet, header_row: int, max_col: int) -> Dict[str, int]:
        """컬럼 매핑 생성"""
        mapping = {}
        
        try:
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=header_row, column=col)
                if cell.value:
                    header_value = str(cell.value).strip().lower()
                    
                    # 각 필드별 키워드 매칭
                    if any(keyword in header_value for keyword in ['사업자등록번호', '사업자번호']):
                        mapping['business_number'] = col
                    elif any(keyword in header_value for keyword in ['상호명', '상호', '업체명', '가맹점명']):
                        mapping['store_name'] = col
                    elif any(keyword in header_value for keyword in ['대표자', '대표자명']):
                        mapping['representative'] = col
                    elif any(keyword in header_value for keyword in ['주소', '사업장주소']):
                        mapping['address'] = col
                    elif any(keyword in header_value for keyword in ['이메일', 'email']):
                        mapping['email'] = col
                    elif any(keyword in header_value for keyword in ['공급가액', '공급가']):
                        mapping['dad_amount'] = col
                    elif any(keyword in header_value for keyword in ['부가세', 'vat']):
                        mapping['mom_amount'] = col
        
        except Exception as e:
            self.logger.warning(f"컬럼 매핑 생성 중 오류: {str(e)}")
        
        return mapping
    
    def _extract_family_from_row(self, sheet, row_num: int, column_mapping: Dict, max_col: int) -> Optional[Dict]:
        """행에서 가족 데이터 추출"""
        try:
            family_data = {}
            
            # 각 필드별 데이터 추출
            for field_name, col_num in column_mapping.items():
                if col_num <= max_col:
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        family_data[field_name] = str(cell.value).strip()
            
            # 필수 필드 확인
            if family_data.get('business_number') and family_data.get('store_name'):
                return family_data
            
        except Exception as e:
            self.logger.debug(f"행 {row_num} 가족 데이터 추출 중 오류: {str(e)}")
        
        return None
    
    def _merge_family_data(self, families: List[Dict]) -> List[Dict]:
        """가족 데이터 통합"""
        try:
            # 사업자번호별로 그룹화
            family_groups = {}
            
            for family in families:
                business_number = family.get('business_number', '').strip()
                if business_number:
                    if business_number not in family_groups:
                        family_groups[business_number] = []
                    family_groups[business_number].append(family)
            
            # 각 그룹을 통합
            merged_families = []
            for business_number, group in family_groups.items():
                if group:
                    merged_family = self._integrate_family_group(group)
                    merged_families.append(merged_family)
            
            return merged_families
            
        except Exception as e:
            self.logger.warning(f"가족 데이터 통합 중 오류: {str(e)}")
            return families
    
    def _integrate_family_group(self, family_group: List[Dict]) -> Dict:
        """가족 그룹 통합"""
        try:
            if not family_group:
                return {}
            
            # 첫 번째 가족을 기본으로 사용
            integrated = family_group[0].copy()
            
            # 다른 가족들의 정보로 누락된 필드 보완
            for family in family_group[1:]:
                for key, value in family.items():
                    if not integrated.get(key) and value:
                        integrated[key] = value
            
            return integrated
            
        except Exception as e:
            self.logger.warning(f"가족 그룹 통합 중 오류: {str(e)}")
            return family_group[0] if family_group else {}


