"""컬럼 매핑 로직 전담 모듈."""

from __future__ import annotations

import logging
from typing import Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from .scoring_utils import score_representative_header

logger = logging.getLogger(__name__)


def map_columns(
    sheet: Worksheet,
    header_row: int,
    required_keywords: Dict[str, List[str]],
) -> Dict[str, int]:
    """필수 컬럼을 헤더 이름 기반으로 매핑한다."""

    column_mapping: Dict[str, int] = {}

    # 검증/완료일자 관련 컬럼 제외를 위한 금지 키워드
    excluded_keywords = ['유효성검사', '외국인제외검증', '완료일자', '검증', '외국인', '제외', '일자별']

    mom_keywords = ['부가세', '세액', 'vat', '세금', '부가세액']
    for col_num in range(1, sheet.max_column + 1):
        raw_value = sheet.cell(header_row, col_num).value
        if raw_value is None:
            continue

        cell_value = str(raw_value).strip().lower()
        
        # 금지 키워드가 포함된 컬럼은 제외
        if any(excluded_keyword in cell_value for excluded_keyword in excluded_keywords):
            logger.debug(
                "검증/완료일자 관련 컬럼 제외: '%s' (컬럼 %d)",
                cell_value,
                col_num,
            )
            continue
        
        if any(keyword in cell_value for keyword in mom_keywords):
            column_mapping['mom_amount'] = col_num
            dad_col = validate_dad_column_before_mom(sheet, col_num, header_row)
            if dad_col is not None:
                column_mapping['dad_amount'] = dad_col
                logger.info(
                    "부가세 앞칸 규칙 적용: 부가세 컬럼 %s 앞칸 %s이 총합계로 확인됨",
                    col_num,
                    dad_col,
                )
            break

    if 'dad_amount' not in column_mapping:
        dad_keywords = ['총', '합계', '금액', '요금', '배달', '총금액', '총합', '배달요금', '총배달요금', '총배달금액']
        for col_num in range(1, sheet.max_column + 1):
            raw_value = sheet.cell(header_row, col_num).value
            if raw_value is None:
                continue
            cell_value = str(raw_value).strip().lower()
            
            # 금지 키워드가 포함된 컬럼은 제외
            if any(excluded_keyword in cell_value for excluded_keyword in excluded_keywords):
                logger.debug(
                    "검증/완료일자 관련 컬럼 제외: '%s' (컬럼 %d)",
                    cell_value,
                    col_num,
                )
                continue
            
            if any(keyword in cell_value for keyword in dad_keywords):
                column_mapping['dad_amount'] = col_num
                break

    business_keywords = ['사업자번호', '사업자등록번호']
    for col_num in range(1, sheet.max_column + 1):
        raw_value = sheet.cell(header_row, col_num).value
        if raw_value is None:
            continue
        cell_value = str(raw_value).strip().lower()
        if any(keyword in cell_value for keyword in business_keywords):
            column_mapping['business_number'] = col_num
            break

    store_keywords = ['상호명', '상호', '업체명', '업체', '매장명', '매장', '점포', '가게', '사업소', '가맹점명']
    for col_num in range(1, sheet.max_column + 1):
        raw_value = sheet.cell(header_row, col_num).value
        if raw_value is None:
            continue
        cell_value = str(raw_value).strip().lower()
        if any(keyword in cell_value for keyword in store_keywords):
            column_mapping['store_name'] = col_num
            break

    representative_fallback_keywords = ['등록자', '등록자명', '작성자', '입력자', '담당자', '담당']
    best_rep_col: Optional[int] = None
    best_rep_header: Optional[str] = None
    best_rep_score = -999
    fallback_rep_col: Optional[int] = None
    fallback_rep_header: Optional[str] = None

    for col_num in range(1, sheet.max_column + 1):
        raw_header_value = sheet.cell(header_row, col_num).value
        if raw_header_value is None:
            continue

        header_value = str(raw_header_value).strip()
        header_lower = header_value.lower()
        score = score_representative_header(header_value)

        if score > best_rep_score:
            best_rep_score = score
            best_rep_col = col_num
            best_rep_header = header_value

        if (
            fallback_rep_col is None
            and any(keyword in header_lower for keyword in representative_fallback_keywords)
        ):
            fallback_rep_col = col_num
            fallback_rep_header = header_value

    if best_rep_col is not None and best_rep_score >= 40:
        column_mapping['representative'] = best_rep_col
        logger.info(
            "대표자 컬럼 매핑: '%s' (점수 %s) -> 열 %s",
            best_rep_header,
            best_rep_score,
            best_rep_col,
        )
    elif best_rep_col is not None and best_rep_score > 0:
        column_mapping['representative'] = best_rep_col
        logger.info(
            "대표자 컬럼 매핑(완전 일치 없음, 점수 %s): '%s' -> 열 %s",
            best_rep_score,
            best_rep_header,
            best_rep_col,
        )
    elif fallback_rep_col is not None:
        column_mapping['representative'] = fallback_rep_col
        logger.info(
            "대표자 컬럼 미발견, 등록자 계열 fallback 사용: '%s' -> 열 %s",
            fallback_rep_header,
            fallback_rep_col,
        )

    address_keywords = ['주소', '사업장주소', '소재지', '사업장', '주소지']
    for col_num in range(1, sheet.max_column + 1):
        raw_value = sheet.cell(header_row, col_num).value
        if raw_value is None:
            continue
        cell_value = str(raw_value).strip().lower()
        if any(keyword in cell_value for keyword in address_keywords):
            column_mapping['address'] = col_num
            break

    email_keywords = ['이메일', 'email', '메일', '사업자이메일']
    for col_num in range(1, sheet.max_column + 1):
        raw_value = sheet.cell(header_row, col_num).value
        if raw_value is None:
            continue
        cell_value = str(raw_value).strip().lower()
        if any(keyword in cell_value for keyword in email_keywords):
            column_mapping['email'] = col_num
            break

    logger.info("컬럼 매핑 결과: %s", column_mapping)
    return column_mapping


def validate_dad_column_before_mom(
    sheet: Worksheet,
    mom_col: int,
    header_row: int,
) -> Optional[int]:
    """부가세 컬럼 앞칸이 총 합계 컬럼인지 10:1 비율로 검증한다."""

    try:
        if mom_col <= 1:
            return None

        dad_col = mom_col - 1
        valid_rows = 0
        total_rows = 0

        for row_num in range(header_row + 1, min(header_row + 20, sheet.max_row + 1)):
            try:
                dad_value = sheet.cell(row_num, dad_col).value
                mom_value = sheet.cell(row_num, mom_col).value

                if not isinstance(dad_value, (int, float)) or not isinstance(mom_value, (int, float)):
                    continue

                if dad_value <= 0 or mom_value <= 0:
                    continue

                total_rows += 1
                ratio = dad_value / mom_value
                if 9.5 <= ratio <= 10.5:
                    valid_rows += 1
            except (ValueError, TypeError):
                continue

        if total_rows > 0 and (valid_rows / total_rows) >= 0.7:
            logger.info(
                "부가세 앞칸 규칙 검증 성공: 컬럼 %s이 총합계로 확인됨 (유효 비율: %s/%s)",
                dad_col,
                valid_rows,
                total_rows,
            )
            return dad_col

        return None

    except Exception as exc:  # pragma: no cover - 방어 로깅
        logger.error("부가세 앞칸 규칙 검증 오류: %s", exc)
        return None

