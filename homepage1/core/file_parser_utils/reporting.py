"""파싱 결과 리포팅/요약 유틸리티."""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ReportingUtils:
    """리포트 및 요약 데이터를 생성하는 헬퍼."""

    def __init__(self, logger: logging.Logger | None = None) -> None:
        self.logger = logger or logging.getLogger(__name__)

    def analyze_data_sections(self, df: pd.DataFrame) -> Dict[str, Any]:
        """데이터 섹션 정보를 분석한다."""

        sections = {
            "header_section": None,
            "data_section": None,
            "summary_section": None,
            "total_rows": len(df),
        }

        try:
            sections["data_section"] = df.to_dict("records")

            if len(df) > 3:
                sections["summary_section"] = df.tail(3).to_dict("records")

            self.logger.info("데이터 섹션 분석 완료: 총 %d행", len(df))
        except Exception as exc:  # pragma: no cover - 보호 로깅
            self.logger.error("데이터 섹션 분석 오류: %s", exc)

        return sections

    def extract_sheet_data(
        self,
        sheet: Worksheet,
        start_row: int,
        headers: List[str],
        *,
        essential_keywords: Optional[List[str]] = None,
    ) -> List[List[Any]]:
        """시트에서 핵심 컬럼만 추출하여 데이터 리스트를 생성한다."""

        data: List[List[Any]] = []

        keywords = essential_keywords or [
            "사업자번호",
            "가맹점명",
            "대표자",
            "주소",
            "이메일",
            "세",
            "금액",
        ]

        essential_columns: List[int] = []
        for index, header in enumerate(headers):
            if any(keyword in header for keyword in keywords):
                essential_columns.append(index)

        max_rows = sheet.max_row

        for row_idx in range(start_row, max_rows + 1):
            row_data = [None] * len(headers)
            has_data = False

            for col_idx in essential_columns:
                cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                if cell_value:
                    row_data[col_idx] = cell_value
                    has_data = True

            if has_data:
                data.append(row_data)

        return data

    def build_priority_sheet_entry(
        self,
        *,
        sheet_name: str,
        headers: List[str],
        data: List[List[Any]],
        header_row: int,
        data_start_row: int,
        priority: Optional[str] = None,
        families: Optional[List[Dict[str, Any]]] = None,
    ) -> Dict[str, Dict[str, Any]]:
        """1순위 시트 정보를 표준 구조로 구성한다."""

        entry = {
            sheet_name: {
                "headers": headers,
                "data": data,
                "header_row": header_row,
                "data_start_row": data_start_row,
            }
        }

        if priority:
            entry[sheet_name]["priority"] = priority

        if families is not None:
            entry[sheet_name]["families"] = families

        return entry

    def collect_sheet_overview(
        self,
        workbook: Workbook,
        *,
        sheet_names: Optional[List[str]] = None,
        header_rows: int = 10,
        data_rows: int = 100,
        column_limit: int = 50,
    ) -> Dict[str, Dict[str, Any]]:
        """워크북의 시트별 요약 정보를 수집한다."""

        overview: Dict[str, Dict[str, Any]] = {}
        target_sheets = sheet_names or list(workbook.sheetnames)

        for sheet_name in target_sheets:
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                self.logger.warning("시트를 찾을 수 없습니다: %s", sheet_name)
                continue

            try:
                sheet_info: Dict[str, Any] = {
                    "sheet_name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "headers": [],
                    "data": [],
                }

                header_limit = min(header_rows, sheet.max_row)
                column_cap = min(column_limit, sheet.max_column)

                for row in range(1, header_limit + 1):
                    row_headers: List[str] = []
                    for col in range(1, column_cap + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        row_headers.append(str(cell_value).strip() if cell_value is not None else "")
                    sheet_info["headers"].append(row_headers)

                data_limit = min(data_rows, sheet.max_row)
                for row in range(1, data_limit + 1):
                    row_data: List[Any] = []
                    for col in range(1, column_cap + 1):
                        row_data.append(sheet.cell(row=row, column=col).value)
                    sheet_info["data"].append(row_data)

                overview[sheet_name] = sheet_info
            except Exception as exc:  # pragma: no cover - 방어적 로깅
                self.logger.warning("시트 '%s' 정보 수집 실패: %s", sheet_name, exc)

        return overview

    def build_preview(self, parsed_data: Dict[str, Any], max_rows: int = 10) -> Dict[str, Any]:
        """미리보기 데이터를 생성한다."""

        if parsed_data.get("parsing_status") != "success":
            return {"error": "파싱 실패"}

        try:
            df: pd.DataFrame = parsed_data["raw_data"]
            preview = df.head(max_rows)
            return {
                "headers": list(df.columns),
                "preview_data": preview.to_dict("records"),
                "total_rows": len(df),
                "file_type": parsed_data.get("file_type"),
            }
        except Exception as exc:  # pragma: no cover - 보호 로깅
            self.logger.error("데이터 미리보기 오류: %s", exc)
            return {"error": f"미리보기 생성 오류: {exc}"}

    def extract_text_content(self, parsed_data: Dict[str, Any]) -> str:
        """텍스트 분석 등을 위한 전체 텍스트 콘텐츠를 생성한다."""

        if parsed_data.get("parsing_status") != "success":
            return ""

        try:
            df: pd.DataFrame = parsed_data["raw_data"]
            texts: list[str] = []

            for _, row in df.iterrows():
                for value in row.values:
                    if pd.notna(value):
                        value_str = str(value).strip()
                        if value_str:
                            texts.append(value_str)

            return " ".join(texts)
        except Exception as exc:  # pragma: no cover - 보호 로깅
            self.logger.error("텍스트 추출 오류: %s", exc)
            return ""


