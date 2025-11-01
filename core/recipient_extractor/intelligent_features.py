"""
지능앱 핵심 기술 모듈

지능앱의 4가지 핵심 기술을 제공합니다:
1. 시트 검열 알고리즘
2. 공급받는자 키워드 매핑 우선순위
3. 이메일 자동 수정 및 검증
4. 수식 대신 결과값으로 읽기
"""

import pandas as pd
import io
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger(__name__)

class IntelligentFeatures:
    """지능앱 핵심 기술 모듈"""
    
    def __init__(self):
        self.logger = logger

    def analyze_sheets_intelligent(self, file_data):
        """지능앱 기술: 시트 검열 알고리즘 (다중 시트 순회 검사, 무제한 헤더 검사, 데이터 밀도 기반 시트 선택)"""
        try:
            sheet_summaries = []
            
            # 다중 시트 순회 검사
            for sheet_name in file_data.sheet_names:
                df = pd.read_excel(file_data, header=None, sheet_name=sheet_name)
                
                # 무제한 헤더 검사
                data_start_row = 0
                possible_headers = []
                
                for row in range(len(df)):
                    row_data = df.iloc[row].astype(str).tolist()
                    if any(keyword in str(cell) for cell in row_data for keyword in ["사업자", "배달", "부가세"]):
                        possible_headers = row_data
                        data_start_row = row
                        # 계속 진행하여 더 아래 후보가 있으면 덮어씀 (choose_last 정책)
                
                # 데이터 밀도 기반 시트 선택
                vat_max = 0
                vat_sum = 0
                
                for col in df.columns:
                    for row in range(data_start_row + 1, len(df)):
                        cell_value = df.iloc[row, col]
                        if isinstance(cell_value, (int, float)) and cell_value > 0:
                            if '부가세' in str(df.iloc[data_start_row, col]) or 'VAT' in str(df.iloc[data_start_row, col]):
                                vat_max = max(vat_max, cell_value)
                                vat_sum += cell_value
                
                sheet_summaries.append({
                    "sheet": sheet_name,
                    "vat_max": vat_max,
                    "vat_sum": vat_sum,
                    "data_start_row": data_start_row,
                    "status": "analyzed"
                })
            
            # VAT 최대값 기준으로 최적 시트 선택 (결정적 정렬)
            if sheet_summaries:
                sheet_summaries.sort(key=lambda x: (x['vat_max'], x.get('sheet', '')), reverse=True)
                best_sheet = sheet_summaries[0]
                self.logger.info(f"지능앱 시트 검열 완료: 최적 시트 '{best_sheet['sheet']}' 선택 (VAT 최대값: {best_sheet['vat_max']})")
                return best_sheet
            else:
                self.logger.warning("지능앱 시트 검열: 분석 가능한 시트가 없습니다")
                return None
                
        except Exception as e:
            self.logger.error(f"지능앱 시트 검열 오류: {str(e)}")
            return None

    def map_recipient_keywords_intelligent(self, df_data, guideline):
        """지능앱 기술: 공급받는자 키워드 매핑 우선순위 (공급받는자 > 수취인 > 거래처 > 일반)"""
        try:
            detected_mapping = {}
            
            if not guideline or 'recipient_keywords' not in guideline:
                self.logger.warning("지능앱 키워드 매핑: 현재 업종 규칙에 recipient_keywords가 없습니다")
                return detected_mapping
            
            recipient_keywords = guideline['recipient_keywords']
            
            for col_idx, col_name in enumerate(df_data.columns):
                col_name_str = str(col_name).strip().lower()
                
                # 공급받는자 사업자번호 (우선순위 높음)
                if "recipient_business_number" not in detected_mapping:
                    recipient_bn_strong = recipient_keywords.get('strong', [])
                    recipient_bn_aux = recipient_keywords.get('aux', [])
                    
                    # 강력한 키워드 우선 매칭
                    for keyword in recipient_bn_strong:
                        if keyword.lower() in col_name_str:
                            detected_mapping["recipient_business_number"] = col_idx
                            self.logger.debug(f"지능앱 키워드 매핑: 강력한 키워드 '{keyword}' 매칭 (컬럼 {col_idx})")
                            break
                    
                    # 강력한 키워드가 없으면 보조 키워드 매칭
                    if "recipient_business_number" not in detected_mapping:
                        for keyword in recipient_bn_aux:
                            if keyword.lower() in col_name_str:
                                detected_mapping["recipient_business_number"] = col_idx
                                self.logger.debug(f"지능앱 키워드 매핑: 보조 키워드 '{keyword}' 매칭 (컬럼 {col_idx})")
                                break
                
                # 상호명 매핑 (배달대행사 특화)
                if "store_name" not in detected_mapping:
                    store_strong = ["공급받는자상호", "수취인상호", "거래처상호"]
                    store_aux = ["상호", "가게명", "매장명", "점포명", "업체명"]
                    
                    for keyword in store_strong:
                        if keyword.lower() in col_name_str:
                            detected_mapping["store_name"] = col_idx
                            self.logger.debug(f"지능앱 키워드 매핑: 상호 강력한 키워드 '{keyword}' 매칭 (컬럼 {col_idx})")
                            break
                    
                    if "store_name" not in detected_mapping:
                        for keyword in store_aux:
                            if keyword.lower() in col_name_str:
                                detected_mapping["store_name"] = col_idx
                                self.logger.debug(f"지능앱 키워드 매핑: 상호 보조 키워드 '{keyword}' 매칭 (컬럼 {col_idx})")
                                break
            
            self.logger.info(f"지능앱 키워드 매핑 완료: {len(detected_mapping)}개 필드 매핑")
            return detected_mapping
            
        except Exception as e:
            self.logger.error(f"지능앱 키워드 매핑 오류: {str(e)}")
            return {}

    def read_sheet_with_data_only_intelligent(self, uploaded_file, sheet_name, header=None):
        """지능앱 기술: 수식 대신 결과값으로 읽기 (openpyxl data_only=True로 시트 읽기)"""
        try:
            # 새 버퍼를 매 호출마다 생성해 포인터 문제 방지
            wb = load_workbook(io.BytesIO(uploaded_file.getbuffer()), data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                # 폴백: 첫 시트
                ws = wb[wb.sheetnames[0]]
                self.logger.warning(f"지능앱 데이터 읽기: 시트 '{sheet_name}'를 찾을 수 없어 첫 번째 시트 '{wb.sheetnames[0]}' 사용")
            else:
                ws = wb[sheet_name]
                self.logger.debug(f"지능앱 데이터 읽기: 시트 '{sheet_name}' 사용")

            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))

            if not rows:
                self.logger.warning(f"지능앱 데이터 읽기: 시트 '{sheet_name}'에 데이터가 없습니다")
                return pd.DataFrame()

            df = pd.DataFrame(rows)
            self.logger.info(f"지능앱 데이터 읽기 완료: 시트 '{sheet_name}', {len(df)}행 {len(df.columns)}열")
            return df
            
        except Exception as e:
            self.logger.error(f"지능앱 데이터 읽기 오류: {str(e)}")
            return pd.DataFrame()

    def detect_template_type(self, file_data: Dict[str, Any]) -> str:
        """템플릿 형태 감지"""
        try:
            # 시트 정보 확인
            sheets_info = file_data.get('sheets_info', {})
            for sheet_name, sheet_info in sheets_info.items():
                headers = sheet_info.get('headers', [])
                
                # 템플릿 특성 키워드 확인
                template_keywords = ["Column_", "엑셀업로드양식", "올바른 예시", "잘못된 예시"]
                structure_keywords = ["홈텍스", "전자세금계산서", "일괄등록"]
                
                header_text = ' '.join(str(h) for h in headers).lower()
                
                # 템플릿 키워드 매칭 확인
                template_matches = sum(1 for keyword in template_keywords if keyword.lower() in header_text)
                structure_matches = sum(1 for keyword in structure_keywords if keyword.lower() in header_text)
                
                if template_matches >= 2 or structure_matches >= 1:
                    self.logger.info(f"템플릿 형태 감지: 시트 '{sheet_name}' (템플릿 매칭: {template_matches}, 구조 매칭: {structure_matches})")
                    return "template_type"
            
            return "normal_type"
            
        except Exception as e:
            self.logger.warning(f"템플릿 감지 오류: {e}")
            return "unknown_type"
