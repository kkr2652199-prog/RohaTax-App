"""
통합 변환 엔진 - 모든 부품을 통합하여 전체 변환 프로세스 실행
핵심기술 절대지침과 공급받는자 지침을 반영하여 홈텍스 템플릿에 기입
"""

import os
import pandas as pd
import openpyxl
import time
from typing import Dict, List, Any, Optional, Tuple
import logging
from pathlib import Path

# 부품들 import
from .file_parser import FileParser
from .recipient_extractor.main_extractor import RecipientExtractor
from .database_manager import db_manager
from .amount_extractor import AmountExtractor
from .template_manager import TemplateManager
from .conversion_core import ConversionCore
from .guideline_manager import GuidelineManager
from .absolute_guideline_loader import get_absolute_guideline_loader

logger = logging.getLogger(__name__)

class ConversionEngine:
    """통합 변환 엔진"""
    
    def __init__(self):
        """변환 엔진 초기화"""
        self.logger = logger
        
        # 부품들 초기화
        self.file_parser = FileParser()
        self.recipient_extractor = RecipientExtractor()
        self.amount_extractor = AmountExtractor()
        self.template_manager = TemplateManager()
        self.conversion_core = ConversionCore()
        self.guideline_manager = GuidelineManager()
        
        # 지능앱 핵심 기술: 절대지침 시스템 초기화
        self.absolute_guideline_loader = get_absolute_guideline_loader()
        
        # 홈텍스 템플릿 컬럼 매핑 (공급받는자 핵심기술 절대지침)
        self.hometax_columns = {
            'K': '사업자등록번호',    # 공급받는자 등록번호
            'M': '상호명',           # 공급받는자 상호
            'N': '대표자명',         # 공급받는자 성명
            'O': '사업장주소',       # 공급받는자 사업장주소
            'R': '이메일',           # 공급받는자 이메일
            'T': '공급가액',         # 공급가액 (1차)
            'U': '부가세',           # 부가세 (1차)
            'AB': '공급가액',        # 공급가액 (2차 - 중복)
            'AC': '부가세'           # 부가세 (2차 - 중복)
        }
    
    def convert_file(self, uploaded_file_path: str, 
                    supplier_info: Dict[str, str], 
                    template_id: str = "hometax_bulk",
                    industry_type: str = "delivery",
                    guidelines: Dict = None,
                    issue_date: str = None,
                    file_name: str = None,
                    user_info: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        전체 변환 프로세스 실행
        
        Args:
            uploaded_file_path: 업로드된 파일 경로
            supplier_info: 공급자 정보 (유저 정보)
            template_id: 템플릿 ID
            
        Returns:
            Dict: 변환 결과
            {
                'success': bool,
                'files': List[str],  # 생성된 파일 경로들
                'total_recipients': int,
                'extraction_summary': Dict,
                'amount_summary': Dict,
                'conversion_log': List[str]
            }
        """
        conversion_log = []
        
        # 📊 상세 통계 수집을 위한 변수 초기화
        start_time = time.time()
        detailed_stats = {
            'total_count': 0,
            'success_rate': 0,
            'rows_processed': 0,
            'files_generated': 0,
            'vat_included_count': 0,
            'vat_zero_count': 0,
            'total_supply_amount': 0,
            'total_tax_amount': 0,
            'email_auto_fixed_count': 0,
            'business_number_auto_fixed_count': 0,
            'perfect_info_count': 0,
            'processing_time': 0,
            'per_second': 0
        }
        
        try:
            conversion_log.append("변환 프로세스 시작")
            
            # ===== 단순화된 변환 프로세스 =====
            self.logger.info("[CONVERSION] 변환 프로세스 시작")
            self.logger.info(f"[CONVERSION] 파일 경로: {uploaded_file_path}")
            self.logger.info(f"[CONVERSION] 사용자 ID: {user_info.get('user_id') if user_info else 'None'}")
            self.logger.info(f"[CONVERSION] 업종 타입: {industry_type or 'delivery'}")
            
            # 기본 사용자 정보 검증만 수행
            if user_info:
                conversion_log.append("기본 사용자 정보 검증")
                if not user_info.get('business_number') or not user_info.get('company_name'):
                    return self._create_error_response("필수 사용자 정보가 누락되었습니다", conversion_log)
                conversion_log.append("기본 사용자 정보 검증 완료")
            
            # ===== 명확한 데이터 전달 구조 =====
            self.logger.info("명확한 데이터 전달 구조 적용 시작")
            
            # 1단계: 파일 파싱
            conversion_log.append("1단계: 파일 파싱 시작")
            self.logger.info("[CONVERSION] 1단계: 파일 파싱 시작")
            parsed_data = self.file_parser.parse_file(uploaded_file_path)
            
            if parsed_data['parsing_status'] != 'success':
                self.logger.error(f"[CONVERSION] 파일 파싱 실패: {parsed_data.get('error_message', '알 수 없는 오류')}")
                return self._create_error_response(
                    f"파일 파싱 실패: {parsed_data.get('error_message', '알 수 없는 오류')}",
                    conversion_log
                )
            
            conversion_log.append(f"파일 파싱 완료: {parsed_data['total_rows']}행")
            self.logger.info(f"[CONVERSION] 파일 파싱 완료: {parsed_data['total_rows']}행")
            
            # 2단계: 업종별 절대지침 적용 (5가지 컬럼 찾기, 6번 7번 컬럼 추출)
            conversion_log.append("2단계: 업종별 절대지침 적용 시작")
            self.logger.info("[CONVERSION] 2단계: 업종별 절대지침 적용 시작")
            selected = self.guideline_manager.select_guideline(industry_type or 'delivery')
            if not self.guideline_manager.is_guideline_ready():
                self.logger.warning(f"[CONVERSION] 지침 미구현 상태로 진행: {selected.get('name') if selected else 'Unknown'} (industry={industry_type})")
            
            # 안전성 검증: selected가 None인 경우 처리
            if selected is None:
                self.logger.error("업종별 지침을 찾을 수 없습니다.")
                return self._create_error_response("업종별 지침을 찾을 수 없습니다.", conversion_log)
            
            # 안전성 검증: selected가 딕셔너리가 아닌 경우 처리
            if not isinstance(selected, dict):
                self.logger.error(f"업종별 지침이 딕셔너리가 아닙니다: {type(selected)}")
                return self._create_error_response("업종별 지침 형식이 올바르지 않습니다.", conversion_log)
            
            # 핵심 수정: 매번 새로운 RecipientExtractor 인스턴스 생성 (상태 초기화)
            recipient_extractor = RecipientExtractor()
            recipient_extractor.set_industry_guideline(industry_type or 'delivery', selected)
            recipients = recipient_extractor.extract_recipients_simple(parsed_data, industry_type or 'delivery')
            
            if not recipients:
                return self._create_error_response(
                    "공급받는자 정보를 추출할 수 없습니다.",
                    conversion_log
                )
            
            # 📊 통계 수집 (recipients에서 _stats 추출)
            if recipients and '_stats' in recipients[0]:
                stats = recipients[0]['_stats']
                detailed_stats.update({
                    'total_count': len(recipients),
                    'success_rate': 100,  # 성공적으로 추출된 경우
                    'rows_processed': stats.get('rows_processed', 0),
                    'vat_included_count': stats.get('vat_included_count', 0),
                    'vat_zero_count': stats.get('vat_zero_count', 0),
                    'total_supply_amount': stats.get('total_supply_amount', 0),
                    'total_tax_amount': stats.get('total_tax_amount', 0),
                    'email_auto_fixed_count': stats.get('email_auto_fixed_count', 0),
                    'business_number_auto_fixed_count': stats.get('business_number_auto_fixed_count', 0),
                    'perfect_info_count': stats.get('perfect_info_count', 0),
                    # 샘플 예시 전달 (있을 때만)
                    'email_auto_fixed_sample_from': stats.get('email_auto_fixed_sample_from'),
                    'email_auto_fixed_sample_to': stats.get('email_auto_fixed_sample_to'),
                    'business_auto_fixed_sample_from': stats.get('business_auto_fixed_sample_from'),
                    'business_auto_fixed_sample_to': stats.get('business_auto_fixed_sample_to')
                })
            
            conversion_log.append(f"업종별 절대지침 적용 완료: {len(recipients)}건")
            
            # 3단계: 공급받는자 통합지침 적용 (템플릿 기입 전용)
            conversion_log.append("3단계: 공급받는자 통합지침 적용 시작")
            self.logger.info("👥 [CONVERSION] 3단계: 공급받는자 통합지침 적용 시작")
            self.logger.info(f"📊 [CONVERSION] 추출 대상 데이터: {len(parsed_data.get('data', []))}행")
            
            # 추출된 데이터를 공급받는자 통합지침에 전달하여 템플릿 기입
            result_files = self._fill_hometax_template_simple(recipients, supplier_info, template_id, issue_date, file_name)
            
            conversion_log.append(f"홈텍스 템플릿 기입 완료: {len(result_files)}개 파일")
            self.logger.info(f"✅ [CONVERSION] 홈텍스 템플릿 기입 완료: {len(result_files)}개 파일")
            
            # 4단계: 결과 요약
            conversion_log.append("4단계: 결과 요약 시작")
            self.logger.info("📋 [CONVERSION] 4단계: 결과 요약 시작")
            extraction_summary = recipient_extractor.get_extraction_summary(recipients)
            
            conversion_log.append("변환 프로세스 완료")
            self.logger.info("🎉 [CONVERSION] 변환 프로세스 완료")
            
            # 📊 최종 통계 완성
            end_time = time.time()
            execution_time = round(end_time - start_time, 2)
            
            # 데이터베이스에 변환 결과 로깅
            db_manager.log_conversion({
                'filename': os.path.basename(uploaded_file_path),
                'file_size': os.path.getsize(uploaded_file_path) if os.path.exists(uploaded_file_path) else 0,
                'recipient_count': len(recipients),
                'success': True,
                'execution_time': execution_time,  # 실행 시간 기록
                'user_id': user_info.get('user_id') if user_info else None
            })
            detailed_stats['processing_time'] = round(end_time - start_time, 2)
            detailed_stats['files_generated'] = len(result_files)
            if detailed_stats['processing_time'] > 0:
                detailed_stats['per_second'] = round(detailed_stats['total_count'] / detailed_stats['processing_time'], 1)
            
            # 📊 상세 통계 계산
            processing_time = time.time() - start_time
            per_second = len(recipients) / processing_time if processing_time > 0 else 0
            
            self.logger.info(f"📊 [CONVERSION] 상세 통계 계산 완료:")
            self.logger.info(f"   - 처리 시간: {processing_time:.2f}초")
            self.logger.info(f"   - 초당 처리 건수: {per_second:.2f}건/초")
            self.logger.info(f"   - 추출된 공급받는자: {len(recipients)}건")
            self.logger.info(f"   - 생성된 파일: {len(result_files)}개")
            
            return {
                'success': True,
                'files': result_files,
                'total_recipients': len(recipients),
                'extraction_summary': extraction_summary,
                'conversion_log': conversion_log,
                'recipients_preview': recipients[:5],  # 처음 5건 미리보기
                'detailed_stats': detailed_stats  # 📊 상세 통계 추가
            }
            
        except Exception as e:
            self.logger.error(f"❌ [CONVERSION] 변환 프로세스 오류: {str(e)}")
            conversion_log.append(f"❌ 변환 프로세스 오류: {str(e)}")
            
            # 에러를 데이터베이스에 로깅
            import traceback
            db_manager.log_error({
                'filename': os.path.basename(uploaded_file_path) if uploaded_file_path else 'unknown',
                'error_type': 'ConversionEngineError',
                'error_message': str(e),
                'stack_trace': traceback.format_exc(),
                'severity': 'ERROR',
                'user_id': user_info.get('user_id') if user_info else None
            })
            
            return self._create_error_response(
                f"변환 프로세스 중 오류 발생: {str(e)}",
                conversion_log
            )
    
    def _filter_valid_data(self, matched_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """업종별 지침에 따른 유효한 데이터 필터링"""
        valid_data = []
        
        # 현재 적용된 업종별 지침 가져오기
        current_guideline = self.recipient_extractor.get_current_guideline()
        min_valid_fields = current_guideline.get('min_valid_fields', 3)
        confidence_threshold = current_guideline.get('confidence_threshold', 0.3)
        
        for data in matched_data:
            # 업종별 필수 항목 확인
            required_fields = ['사업자등록번호', '상호', '대표명', '사업장주소', '사업자이메일']
            
            valid_fields = sum(1 for field in required_fields 
                              if data.get(field, '').strip())
            
            # 업종별 최소 필드 수 확인
            if valid_fields >= min_valid_fields:
                # 신뢰도 확인 (있는 경우)
                if 'confidence' in data:
                    if data['confidence'] >= confidence_threshold:
                        valid_data.append(data)
                else:
                    # 신뢰도가 없으면 그대로 통과
                    valid_data.append(data)
        
        self.logger.info(f"유효 데이터 필터링: {len(matched_data)}건 → {len(valid_data)}건 (업종: {current_guideline.get('industry', 'unknown')}, 최소필드: {min_valid_fields}, 신뢰도: {confidence_threshold})")
        return valid_data
    
    def _fill_hometax_template_simple(self, recipients: List[Dict[str, Any]], 
                                     supplier_info: Dict[str, str], 
                                     template_id: str,
                                     issue_date: str = None,
                                     file_name: str = None) -> List[str]:
        """
        단순한 홈텍스 템플릿 기입 (공급받는자 통합지침 적용)
        """
        try:
            # 파일 분할 (50개씩)
            file_count = self.conversion_core.calculate_file_count(len(recipients))
            result_files = []
            
            for file_index in range(file_count):
                # 해당 파일의 데이터 범위 계산
                start_idx, end_idx = self.conversion_core.get_supplier_range(file_index)
                file_data = recipients[start_idx:end_idx]
                
                # 홈텍스 템플릿 생성
                template_path = self._create_hometax_file_simple(
                    data=file_data, 
                    supplier_info=supplier_info, 
                    file_index=file_index,
                    template_id=template_id,
                    issue_date=issue_date,
                    file_name=file_name
                )
                result_files.append(template_path)
            
            return result_files
            
        except Exception as e:
            self.logger.error(f"단순 홈텍스 템플릿 기입 오류: {str(e)}")
            raise
    
    def _create_hometax_file_simple(self, data: List[Dict[str, Any]], 
                                   supplier_info: Dict[str, str], 
                                   file_index: int,
                                   template_id: str,
                                   issue_date: str = None,
                                   file_name: str = None) -> str:
        """단순한 홈텍스 파일 생성 (공급받는자 통합지침 적용)"""
        try:
            wb = None
            ws = None
            # 공식 템플릿 로드
            template_path = self.template_manager.get_template_path(template_id)
            if not template_path:
                raise RuntimeError(f"공식 템플릿을 찾을 수 없습니다: template_id={template_id}")
            wb = openpyxl.load_workbook(template_path)
            template_info = self.template_manager.get_template_info(template_id) or {}
            sheet_name = '엑셀업로드양식'
            if sheet_name not in wb.sheetnames:
                fallback_sheet = template_info.get('sheet_name') or wb.active.title
                if fallback_sheet in wb.sheetnames:
                    sheet_name = fallback_sheet
                else:
                    raise RuntimeError(f"템플릿 시트를 찾을 수 없습니다: 요구='엑셀업로드양식', 보유={wb.sheetnames}")
            ws = wb[sheet_name]
            
            # 실제 추출 건수만큼 공급자/절대값/데이터를 기입
            data_len = len(data)
            self.logger.info(f"단순 템플릿 기입 시작 - 추출 건수: {data_len}")
            if data_len == 0:
                self.logger.warning("유효한 공급받는자 데이터가 0건입니다.")
            
            # 공급자 정보: 최소 1행은 기입하여 템플릿 유효성을 보장
            supplier_rows = max(1, data_len)
            self._set_supplier_info(ws, supplier_info, num_rows=supplier_rows, issue_date=issue_date)
            
            # 공급받는자 데이터 기입 (7행부터)
            self._fill_recipient_data_simple(ws, data)
            
            # 절대값 규칙 적용 (공급자 기입 범위와 동일)
            self._apply_absolute_values(ws, num_rows=supplier_rows)
            
            # 파일 저장
            output_dir = Path("output")
            output_dir.mkdir(exist_ok=True)
            
            # 파일명 생성: 사용자 입력 파일명 사용하거나 기본값 사용
            if file_name:
                # 사용자 입력 파일명에서 확장자 제거 후 인덱스 추가
                base_name = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
                filename = f"{base_name}_{file_index + 1:02d}.xlsx"
            else:
                # 기본 파일명 사용
                filename = f"hometax_bulk_{file_index + 1:02d}.xlsx"
            
            file_path = output_dir / filename
            
            wb.save(file_path)
            self.logger.info(f"단순 홈텍스 파일 생성 완료: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            self.logger.error(f"단순 홈텍스 파일 생성 오류: {str(e)}")
            raise
    
    def _fill_recipient_data_simple(self, ws, data: List[Dict[str, Any]]):
        """단순한 공급받는자 데이터 기입 (공급받는자 통합지침 적용)"""
        start_row = 7
        
        for i, recipient in enumerate(data):
            current_row = start_row + i
            
            # 공급받는자 통합지침: 표준화된 필드명 사용
            biz_no = recipient.get('사업자등록번호', '')
            store_name = recipient.get('상호', '')
            rep_name = recipient.get('대표명', '')
            address = recipient.get('사업장주소', '')
            email = recipient.get('사업자이메일', '')
            supply_amount = recipient.get('공급가액', 0) or 0
            vat_amount = recipient.get('부가세', 0) or 0
            
            # 셀 기입 (홈텍스 템플릿 표준)
            ws[f'K{current_row}'] = biz_no
            ws[f'M{current_row}'] = store_name
            ws[f'N{current_row}'] = rep_name
            ws[f'O{current_row}'] = address
            ws[f'R{current_row}'] = email
            
            ws[f'T{current_row}'] = supply_amount
            ws[f'U{current_row}'] = vat_amount
            ws[f'AB{current_row}'] = supply_amount
            ws[f'AC{current_row}'] = vat_amount
            
            # 순서 보장 로깅 (처음 5건과 특정 순서)
            if i < 5 or i in [49, 99, 149]:  # 1-5번째, 50번째, 100번째, 150번째
                self.logger.info(f"고객 {i+1}번째 기입: {store_name} → {current_row}행 (공식: {start_row} + {i})")
            
            self.logger.debug(f"공급받는자 데이터 기입: 행 {current_row}, 상호: {store_name}, 금액: {supply_amount}/{vat_amount}")
    
    def _fill_hometax_template(self, valid_data: List[Dict[str, Any]], 
                              supplier_info: Dict[str, str], 
                              template_id: str,
                              issue_date: str = None,
                              file_name: str = None) -> List[str]:
        """
        홈텍스 템플릿에 데이터 기입
        공급받는자 핵심기술 절대지침에 따라 컬럼 매핑
        """
        try:
            # 파일 분할 (50개씩)
            file_count = self.conversion_core.calculate_file_count(len(valid_data))
            result_files = []
            
            for file_index in range(file_count):
                # 해당 파일의 데이터 범위 계산
                start_idx, end_idx = self.conversion_core.get_supplier_range(file_index)
                file_data = valid_data[start_idx:end_idx]
                
                # 홈텍스 템플릿 생성
                template_path = self._create_hometax_file(
                    data=file_data, 
                    supplier_info=supplier_info, 
                    file_index=file_index,
                    template_id=template_id,
                    issue_date=issue_date,
                    file_name=file_name
                )
                result_files.append(template_path)
            
            return result_files
            
        except Exception as e:
            self.logger.error(f"홈텍스 템플릿 기입 오류: {str(e)}")
            raise
    
    def _create_hometax_file(self, data: List[Dict[str, Any]], 
                            supplier_info: Dict[str, str], 
                            file_index: int,
                            template_id: str,
                            issue_date: str = None,
                            file_name: str = None) -> str:
        """개별 홈텍스 파일 생성
        - 가능하면 공식 홈텍스 템플릿을 로드하여 그 안에 7행부터 기입
        - 템플릿을 찾지 못하면 최소 스켈레톤을 생성
        """
        try:
            wb = None
            ws = None
            # 1) 공식 템플릿 로드 (강제)
            template_path = self.template_manager.get_template_path(template_id)
            if not template_path:
                raise RuntimeError(f"공식 템플릿을 찾을 수 없습니다: template_id={template_id}")
            wb = openpyxl.load_workbook(template_path)
            template_info = self.template_manager.get_template_info(template_id) or {}
            sheet_name = '엑셀업로드양식'
            if sheet_name not in wb.sheetnames:
                # 설정에 지정된 시트명 시도
                fallback_sheet = template_info.get('sheet_name') or wb.active.title
                if fallback_sheet in wb.sheetnames:
                    sheet_name = fallback_sheet
                else:
                    raise RuntimeError(f"템플릿 시트를 찾을 수 없습니다: 요구='엑셀업로드양식', 보유={wb.sheetnames}")
            ws = wb[sheet_name]
            
            # 실제 추출 건수만큼 공급자/절대값/데이터를 기입
            data_len = len(data)
            self.logger.info(f"템플릿 기입 시작 - 추출 건수: {data_len}")
            if data_len == 0:
                self.logger.warning("유효한 공급받는자 데이터가 0건입니다. 지침/매핑을 확인하세요.")

            # 공급자 정보: 최소 1행은 기입하여 템플릿 유효성을 보장
            supplier_rows = max(1, data_len)
            self._set_supplier_info(ws, supplier_info, num_rows=supplier_rows, issue_date=issue_date)
            
            # 공급받는자 데이터 기입 (7행부터)
            self._fill_recipient_data(ws, data)
            
            # 절대값 규칙 적용 (공급자 기입 범위와 동일)
            self._apply_absolute_values(ws, num_rows=supplier_rows)
            
            # 파일 저장
            output_dir = Path("output")
            output_dir.mkdir(exist_ok=True)
            
            # 파일명 생성: 사용자 입력 파일명 사용하거나 기본값 사용
            if file_name:
                # 사용자 입력 파일명에서 확장자 제거 후 인덱스 추가
                base_name = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
                filename = f"{base_name}_{file_index + 1:02d}.xlsx"
            else:
                # 기본 파일명 사용
                filename = f"hometax_bulk_{file_index + 1:02d}.xlsx"
            
            file_path = output_dir / filename
            
            wb.save(file_path)
            self.logger.info(f"홈텍스 파일 생성 완료: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            self.logger.error(f"홈텍스 파일 생성 오류: {str(e)}")
            raise
    
    def _set_hometax_headers(self, ws):
        """홈텍스 헤더 설정 (6행)"""
        headers = {
            'K': '사업자등록번호',
            'M': '상호명', 
            'N': '대표자명',
            'O': '사업장주소',
            'R': '이메일',
            'T': '공급가액',
            'U': '부가세',
            'AB': '공급가액',
            'AC': '부가세'
        }
        
        for col, header in headers.items():
            ws[f'{col}6'] = header
    
    def _set_supplier_info(self, ws, supplier_info: Dict[str, str], num_rows: int = 1, issue_date: str = None):
        """공급자 정보 설정 (7행부터 num_rows만큼 반복 기입)
        - 요구사항: 공급자 정보는 추출 건수만큼 동일하게 기입
        - BG 열 절대값 '01'도 함께 기록
        """
        start_row = 7
        end_row = start_row + max(0, num_rows - 1)

        for row_num in range(start_row, end_row + 1):
            # 세금일자 포맷 변환 (YYYY-MM-DD → YYYYMMDD)
            tax_date = '20251001'  # 기본값
            if issue_date:
                try:
                    # ISO 형식 (YYYY-MM-DD)을 YYYYMMDD로 변환
                    from datetime import datetime
                    date_obj = datetime.fromisoformat(issue_date)
                    tax_date = date_obj.strftime('%Y%m%d')
                except:
                    # 변환 실패 시 기본값 사용
                    tax_date = '20251001'
            
            supplier_mapping = {
                'A': '01',  # 전자세금계산서 종류 (절대값)
                'B': tax_date,  # 작성일자 (유저 지정)
                'C': supplier_info.get('supplier_business_number', ''),
                'D': '',  # 공급자 종사업장번호
                'E': supplier_info.get('supplier_name', ''),
                'F': supplier_info.get('supplier_representative', ''),
                'G': supplier_info.get('supplier_address', ''),
                'H': supplier_info.get('supplier_business_type', ''),
                'I': supplier_info.get('supplier_business_category', ''),
                'J': supplier_info.get('supplier_email', ''),
                'W': '30',  # 일자1 (절대값)
                'BG': '01'  # 절대값 (요구 반영)
            }

            for col, value in supplier_mapping.items():
                ws[f'{col}{row_num}'] = value
                self.logger.debug(f"공급자 정보 기입: {col}{row_num} = {value}")
    
    def _fill_recipient_data(self, ws, data: List[Dict[str, Any]]):
        """공급받는자 데이터 기입 (7행부터) - 다양한 키(지침별 별칭) 지원"""
        def get_value(record: Dict[str, Any], aliases: List[str], default: Any = ""):
            for key in aliases:
                if key in record and record.get(key) not in [None, ""]:
                    return record.get(key)
            return default

        # 1) 템플릿 전용 집계층: 부가세>0인 행만 사업자번호별 합산
        grouped: Dict[str, Dict[str, Any]] = {}
        for rec in data:
            biz_no = get_value(rec, ['사업자등록번호', '등록번호', 'buyer_biz_no', 'business_number']).strip()
            # 부가세가 0/없음이면 제외 (2순위 정책)
            try:
                vat_val = float(get_value(rec, ['부가세', '세액', 'vat', 'tax_amount'], 0) or 0)
            except Exception:
                vat_val = 0.0
            if vat_val <= 0:
                continue

            try:
                supply_val = float(get_value(rec, ['공급가액', '공급가액(1차)', 'supply_amount'], 0) or 0)
            except Exception:
                supply_val = 0.0

            if biz_no not in grouped:
                grouped[biz_no] = {
                    '사업자등록번호': biz_no,
                    '상호': get_value(rec, ['상호', '상호명', '업체명', '가맹점명', 'store_name', 'buyer_name']),
                    '대표명': get_value(rec, ['대표명', '대표자', '대표자명', 'owner', 'representative']),
                    '사업장주소': get_value(rec, ['사업장주소', '주소', '사업장 주소', 'address']),
                    '사업자이메일': get_value(rec, ['사업자이메일', '이메일', 'email', 'email1']),
                    '공급가액': 0.0,
                    '부가세': 0.0,
                }
            grouped[biz_no]['공급가액'] = float(grouped[biz_no]['공급가액']) + supply_val
            grouped[biz_no]['부가세'] = float(grouped[biz_no]['부가세']) + vat_val

        # 집계 결과 리스트 (기존 순서 무관)
        aggregated_data = []
        for g in grouped.values():
            g['요금합계'] = float(g['공급가액']) + float(g['부가세'])
            aggregated_data.append(g)

        start_row = 7

        # 2) 집계된 데이터만 템플릿에 기입
        for i, recipient in enumerate(aggregated_data):
            current_row = start_row + i

            # 필드 별칭 정의 (여러 지침/추출기 결과 호환)
            biz_no = get_value(recipient, ['사업자등록번호', '등록번호', 'buyer_biz_no', 'business_number'])
            store_name = get_value(recipient, ['상호', '상호명', '업체명', '가맹점명', 'store_name', 'buyer_name'])
            rep_name = get_value(recipient, ['대표명', '대표자', '대표자명', 'owner', 'representative'])
            address = get_value(recipient, ['사업장주소', '주소', '사업장 주소', 'address'])
            email = get_value(recipient, ['사업자이메일', '이메일', 'email', 'email1'])

            # 금액 필드 별칭 및 계산
            supply_amount = get_value(recipient, ['공급가액', '공급가액(1차)', 'supply_amount'], 0) or 0
            vat_amount = get_value(recipient, ['부가세', '세액', 'vat', 'tax_amount'], 0) or 0
            total_fee = get_value(recipient, ['요금합계', '총액', '합계', 'total_amount'], 0) or 0

            if not supply_amount and total_fee and vat_amount is not None:
                try:
                    supply_amount = float(total_fee) - float(vat_amount)
                except Exception:
                    pass

            # 셀 기입
            ws[f'K{current_row}'] = biz_no
            ws[f'M{current_row}'] = store_name
            ws[f'N{current_row}'] = rep_name
            ws[f'O{current_row}'] = address
            ws[f'R{current_row}'] = email

            ws[f'T{current_row}'] = supply_amount
            ws[f'U{current_row}'] = vat_amount
            ws[f'AB{current_row}'] = supply_amount
            ws[f'AC{current_row}'] = vat_amount

            # 순서 보장 로깅 (처음 5건과 특정 순서)
            if i < 5 or i in [49, 99, 149]:  # 1-5번째, 50번째, 100번째, 150번째
                self.logger.info(f"고객 {i+1}번째 기입: {store_name} → {current_row}행 (공식: {start_row} + {i})")

            self.logger.debug(f"공급받는자 데이터 기입: 행 {current_row}, 상호: {store_name}, 금액: {supply_amount}/{vat_amount}")
    
    def _apply_absolute_values(self, ws, num_rows: int = None):
        """절대값 규칙 적용 (데이터 건수만큼만)
        - A열: '01'
        - W열: '30'
        - BG열: '01' (요구 반영)
        """
        try:
            start_row = 7
            if num_rows is None:
                end_row = ws.max_row
            else:
                end_row = start_row + max(0, num_rows - 1)

            for row in range(start_row, end_row + 1):
                ws[f'A{row}'] = '01'
                ws[f'W{row}'] = '30'
                ws[f'BG{row}'] = '01'

            self.logger.debug(f"절대값 규칙 적용 완료: rows {start_row}-{end_row}")
        except Exception as e:
            self.logger.warning(f"절대값 규칙 적용 오류: {str(e)}")
    
    def _create_error_response(self, error_message: str, conversion_log: List[str]) -> Dict[str, Any]:
        """오류 응답 생성"""
        return {
            'success': False,
            'error_message': error_message,
            'files': [],
            'total_recipients': 0,
            'extraction_summary': {},
            'amount_summary': {},
            'conversion_log': conversion_log
        }
    
    def get_conversion_status(self, conversion_result: Dict[str, Any]) -> Dict[str, Any]:
        """변환 결과 상태 요약"""
        if not conversion_result['success']:
            return {
                'status': 'failed',
                'message': conversion_result.get('error_message', '변환 실패'),
                'files_count': 0,
                'recipients_count': 0
            }
        
        return {
            'status': 'success',
            'message': '변환 완료',
            'files_count': len(conversion_result.get('files', [])),
            'recipients_count': conversion_result.get('total_recipients', 0),
            'extraction_rate': conversion_result.get('extraction_summary', {}).get('extraction_rate', 0),
            'total_amount': conversion_result.get('amount_summary', {}).get('total_amount', 0)
        }

# 테스트용 함수
def test_conversion_engine():
    """ConversionEngine 테스트"""
    engine = ConversionEngine()
    
    # 테스트 데이터
    test_data = {
        '가맹점명': ['신전떡볶이', '맘스터치', '피자헛'],
        '사업자번호': ['123-45-67890', '234-56-78901', '345-67-89012'],
        '대표자명': ['홍길동', '김철수', '이영희'],
        '주소': ['서울시 강남구 테헤란로 123', '부산시 해운대구 센텀로 456', '대구시 수성구 동대구로 789'],
        '이메일': ['hong@shinjeon.com', 'kim@moms.com', 'lee@pizza.com'],
        '요금합계': ['50,000원', '75,000원', '60,000원'],
        '부가세': ['5,000원', '7,500원', '6,000원']
    }
    
    df = pd.DataFrame(test_data)
    
    # 공급자 정보
    supplier_info = {
        'supplier_name': '테스트공급자',
        'supplier_representative': '테스트대표',
        'supplier_business_number': '999-99-99999',
        'supplier_address': '서울시 테스트구 테스트로 123',
        'supplier_email': 'test@supplier.com'
    }
    
    # 파싱 결과 시뮬레이션
    parsed_data = {
        'parsing_status': 'success',
        'raw_data': df
    }
    
    # 공급받는자 정보 추출
    recipients = engine.recipient_extractor.extract_recipients(parsed_data)
    amounts = engine.amount_extractor.extract_amounts(parsed_data)
    matched_data = engine.amount_extractor.match_amounts_with_recipients(recipients, amounts)
    
    print("매칭된 데이터:")
    for i, data in enumerate(matched_data, 1):
        print(f"{i}. {data}")
    
    # 변환 상태 확인
    status = engine.get_conversion_status({'success': True, 'files': ['test.xlsx'], 'total_recipients': len(matched_data)})
    print(f"\n변환 상태: {status}")

if __name__ == "__main__":
    test_conversion_engine()









