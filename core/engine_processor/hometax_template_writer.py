"""홈택스 템플릿 작성 전용 모듈."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


class HometaxTemplateWriter:
    """홈택스 템플릿 작성 로직을 담당하는 보조 클래스."""

    def __init__(self, template_manager, conversion_core, logger: Optional[logging.Logger] = None) -> None:
        self.template_manager = template_manager
        self.conversion_core = conversion_core
        self.logger = logger or logging.getLogger(__name__)

    # ------------------------------------------------------------------
    # 퍼블릭 API
    # ------------------------------------------------------------------
    def fill_templates_simple(
        self,
        recipients: List[Dict[str, Any]],
        supplier_info: Dict[str, str],
        template_id: str,
        issue_date: Optional[str] = None,
        file_name: Optional[str] = None,
    ) -> List[str]:
        """단순 공급받는자 통합지침 데이터를 템플릿에 기입한다."""

        file_count = self.conversion_core.calculate_file_count(len(recipients))
        result_files: List[str] = []

        for file_index in range(file_count):
            start_idx, end_idx = self.conversion_core.get_supplier_range(file_index)
            file_data = recipients[start_idx:end_idx]

            output_path = self._create_hometax_file_simple(
                data=file_data,
                supplier_info=supplier_info,
                file_index=file_index,
                template_id=template_id,
                issue_date=issue_date,
                file_name=file_name,
            )
            result_files.append(output_path)

        return result_files

    def fill_templates(
        self,
        valid_data: List[Dict[str, Any]],
        supplier_info: Dict[str, str],
        template_id: str,
        issue_date: Optional[str] = None,
        file_name: Optional[str] = None,
    ) -> List[str]:
        """공급받는자 핵심지침 데이터를 템플릿에 기입한다."""

        file_count = self.conversion_core.calculate_file_count(len(valid_data))
        result_files: List[str] = []

        for file_index in range(file_count):
            start_idx, end_idx = self.conversion_core.get_supplier_range(file_index)
            file_data = valid_data[start_idx:end_idx]

            output_path = self._create_hometax_file(
                data=file_data,
                supplier_info=supplier_info,
                file_index=file_index,
                template_id=template_id,
                issue_date=issue_date,
                file_name=file_name,
            )
            result_files.append(output_path)

        return result_files

    # ------------------------------------------------------------------
    # 내부 구현: 템플릿 생성 및 저장
    # ------------------------------------------------------------------
    def _create_hometax_file_simple(
        self,
        data: List[Dict[str, Any]],
        supplier_info: Dict[str, str],
        file_index: int,
        template_id: str,
        issue_date: Optional[str] = None,
        file_name: Optional[str] = None,
    ) -> str:
        workbook, worksheet = self._load_template(template_id)

        data_len = len(data)
        self.logger.info("단순 템플릿 기입 시작 - 추출 건수: %s", data_len)
        if data_len == 0:
            self.logger.warning("유효한 공급받는자 데이터가 0건입니다.")

        supplier_rows = max(1, data_len)
        self._set_supplier_info(worksheet, supplier_info, num_rows=supplier_rows, issue_date=issue_date)
        self._fill_recipient_data_simple(worksheet, data)
        self._apply_absolute_values(worksheet, num_rows=supplier_rows)

        return self._save_workbook(workbook, file_index, file_name)

    def _create_hometax_file(
        self,
        data: List[Dict[str, Any]],
        supplier_info: Dict[str, str],
        file_index: int,
        template_id: str,
        issue_date: Optional[str] = None,
        file_name: Optional[str] = None,
    ) -> str:
        workbook, worksheet = self._load_template(template_id)

        data_len = len(data)
        self.logger.info("템플릿 기입 시작 - 추출 건수: %s", data_len)
        if data_len == 0:
            self.logger.warning("유효한 공급받는자 데이터가 0건입니다. 지침/매핑을 확인하세요.")

        supplier_rows = max(1, data_len)
        self._set_supplier_info(worksheet, supplier_info, num_rows=supplier_rows, issue_date=issue_date)
        self._fill_recipient_data(worksheet, data)
        self._apply_absolute_values(worksheet, num_rows=supplier_rows)

        return self._save_workbook(workbook, file_index, file_name)

    def _load_template(self, template_id: str):
        template_path = self.template_manager.get_template_path(template_id)
        if not template_path:
            raise RuntimeError(f"공식 템플릿을 찾을 수 없습니다: template_id={template_id}")

        workbook = openpyxl.load_workbook(template_path)
        template_info = self.template_manager.get_template_info(template_id) or {}

        sheet_name = "엑셀업로드양식"
        if sheet_name not in workbook.sheetnames:
            fallback_sheet = template_info.get("sheet_name") or workbook.active.title
            if fallback_sheet in workbook.sheetnames:
                sheet_name = fallback_sheet
            else:
                raise RuntimeError(
                    f"템플릿 시트를 찾을 수 없습니다: 요구='엑셀업로드양식', 보유={workbook.sheetnames}"
                )

        worksheet = workbook[sheet_name]
        return workbook, worksheet

    def _save_workbook(self, workbook, file_index: int, file_name: Optional[str]) -> str:
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        if file_name:
            base_name = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
            filename = f"{base_name}_{file_index + 1:02d}.xlsx"
        else:
            filename = f"hometax_bulk_{file_index + 1:02d}.xlsx"

        file_path = output_dir / filename
        workbook.save(file_path)
        self.logger.info("홈텍스 파일 생성 완료: %s", file_path)
        return str(file_path)

    # ------------------------------------------------------------------
    # 데이터 기입 보조 메서드
    # ------------------------------------------------------------------
    def _set_supplier_info(
        self,
        worksheet,
        supplier_info: Dict[str, str],
        num_rows: int = 1,
        issue_date: Optional[str] = None,
    ) -> None:
        start_row = 7
        end_row = start_row + max(0, num_rows - 1)

        for row_num in range(start_row, end_row + 1):
            tax_date = "20251001"
            if issue_date:
                try:
                    from datetime import datetime

                    date_obj = datetime.fromisoformat(issue_date)
                    tax_date = date_obj.strftime("%Y%m%d")
                except Exception:  # pragma: no cover - 변환 실패 시 기본값 유지
                    tax_date = "20251001"

            supplier_mapping = {
                "A": "01",
                "B": tax_date,
                "C": supplier_info.get("supplier_business_number", ""),
                "D": "",
                "E": supplier_info.get("supplier_name", ""),
                "F": supplier_info.get("supplier_representative", ""),
                "G": supplier_info.get("supplier_address", ""),
                "H": supplier_info.get("supplier_business_type", ""),
                "I": supplier_info.get("supplier_business_category", ""),
                "J": supplier_info.get("supplier_email", ""),
                "W": "30",
                "BG": "01",
            }

            for col, value in supplier_mapping.items():
                worksheet[f"{col}{row_num}"] = value
                self.logger.debug("공급자 정보 기입: %s%s = %s", col, row_num, value)

    def _fill_recipient_data_simple(self, worksheet, data: List[Dict[str, Any]]) -> None:
        start_row = 7

        for i, recipient in enumerate(data):
            current_row = start_row + i

            biz_no = recipient.get("사업자등록번호", "")
            store_name = recipient.get("상호", "")
            rep_name = recipient.get("대표명", "")
            address = recipient.get("사업장주소", "")
            email = recipient.get("사업자이메일", "")
            supply_amount = recipient.get("공급가액", 0) or 0
            vat_amount = recipient.get("부가세", 0) or 0

            worksheet[f"K{current_row}"] = biz_no
            worksheet[f"M{current_row}"] = store_name
            worksheet[f"N{current_row}"] = rep_name
            worksheet[f"O{current_row}"] = address
            worksheet[f"R{current_row}"] = email

            worksheet[f"T{current_row}"] = supply_amount
            worksheet[f"U{current_row}"] = vat_amount
            worksheet[f"AB{current_row}"] = supply_amount
            worksheet[f"AC{current_row}"] = vat_amount

            if i < 5 or i in [49, 99, 149]:
                self.logger.info("고객 %s번째 기입: %s → %s행", i + 1, store_name, current_row)

            self.logger.debug(
                "공급받는자 데이터 기입(단순): 행 %s, 상호: %s, 금액: %s/%s",
                current_row,
                store_name,
                supply_amount,
                vat_amount,
            )

    def _fill_recipient_data(self, worksheet, data: List[Dict[str, Any]]) -> None:
        def get_value(record: Dict[str, Any], aliases: List[str], default: Any = ""):
            for key in aliases:
                if key in record and record.get(key) not in [None, ""]:
                    return record.get(key)
            return default

        grouped: Dict[str, Dict[str, Any]] = {}
        for rec in data:
            biz_no = get_value(rec, ["사업자등록번호", "등록번호", "buyer_biz_no", "business_number"]).strip()
            try:
                vat_val = float(get_value(rec, ["부가세", "세액", "vat", "tax_amount"], 0) or 0)
            except Exception:
                vat_val = 0.0
            if vat_val <= 0:
                continue

            try:
                supply_val = float(get_value(rec, ["공급가액", "공급가액(1차)", "supply_amount"], 0) or 0)
            except Exception:
                supply_val = 0.0

            if biz_no not in grouped:
                grouped[biz_no] = {
                    "사업자등록번호": biz_no,
                    "상호": get_value(rec, ["상호", "상호명", "업체명", "가맹점명", "store_name", "buyer_name"]),
                    "대표명": get_value(rec, ["대표명", "대표자", "대표자명", "owner", "representative"]),
                    "사업장주소": get_value(rec, ["사업장주소", "주소", "사업장 주소", "address"]),
                    "사업자이메일": get_value(rec, ["사업자이메일", "이메일", "email", "email1"]),
                    "공급가액": 0.0,
                    "부가세": 0.0,
                }

            grouped[biz_no]["공급가액"] = float(grouped[biz_no]["공급가액"]) + supply_val
            grouped[biz_no]["부가세"] = float(grouped[biz_no]["부가세"]) + vat_val

        aggregated_data = []
        for group in grouped.values():
            group["요금합계"] = float(group["공급가액"]) + float(group["부가세"])
            aggregated_data.append(group)

        start_row = 7
        for i, recipient in enumerate(aggregated_data):
            current_row = start_row + i

            biz_no = get_value(recipient, ["사업자등록번호", "등록번호", "buyer_biz_no", "business_number"])
            store_name = get_value(recipient, ["상호", "상호명", "업체명", "가맹점명", "store_name", "buyer_name"])
            rep_name = get_value(recipient, ["대표명", "대표자", "대표자명", "owner", "representative"])
            address = get_value(recipient, ["사업장주소", "주소", "사업장 주소", "address"])
            email = get_value(recipient, ["사업자이메일", "이메일", "email", "email1"])

            supply_amount = get_value(recipient, ["공급가액", "공급가액(1차)", "supply_amount"], 0) or 0
            vat_amount = get_value(recipient, ["부가세", "세액", "vat", "tax_amount"], 0) or 0
            total_fee = get_value(recipient, ["요금합계", "총액", "합계", "total_amount"], 0) or 0

            if not supply_amount and total_fee and vat_amount is not None:
                try:
                    supply_amount = float(total_fee) - float(vat_amount)
                except Exception:
                    pass

            worksheet[f"K{current_row}"] = biz_no
            worksheet[f"M{current_row}"] = store_name
            worksheet[f"N{current_row}"] = rep_name
            worksheet[f"O{current_row}"] = address
            worksheet[f"R{current_row}"] = email

            worksheet[f"T{current_row}"] = supply_amount
            worksheet[f"U{current_row}"] = vat_amount
            worksheet[f"AB{current_row}"] = supply_amount
            worksheet[f"AC{current_row}"] = vat_amount

            if i < 5 or i in [49, 99, 149]:
                self.logger.info("고객 %s번째 기입: %s → %s행", i + 1, store_name, current_row)

            self.logger.debug(
                "공급받는자 데이터 기입: 행 %s, 상호: %s, 금액: %s/%s",
                current_row,
                store_name,
                supply_amount,
                vat_amount,
            )

    def _apply_absolute_values(self, worksheet, num_rows: Optional[int] = None) -> None:
        start_row = 7
        if num_rows is None:
            end_row = worksheet.max_row
        else:
            end_row = start_row + max(0, num_rows - 1)

        for row in range(start_row, end_row + 1):
            worksheet[f"A{row}"] = "01"
            worksheet[f"W{row}"] = "30"
            worksheet[f"BG{row}"] = "01"

        self.logger.debug("절대값 규칙 적용 완료: rows %s-%s", start_row, end_row)

