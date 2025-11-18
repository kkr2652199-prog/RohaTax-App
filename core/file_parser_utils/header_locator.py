"""헤더 및 시트 감지 로직 전담 모듈."""

from __future__ import annotations

import logging
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from .config_builder import build_forbidden_keywords_map, get_scoring_config
from .number_parser import default_number_parser
from .column_mapper import map_columns, validate_dad_column_before_mom
from .header_detector import (
    detect_header_row,
    detect_csv_header_row,
    detect_title_rows,
    analyze_text_pattern,
    is_title_pattern,
    find_actual_data_range,
    get_actual_data_range,
)
from .scoring_utils import (
    score_representative_header,
    count_matched_fields,
    calculate_data_density,
    evaluate_data_quality,
    calculate_csv_data_density,
    count_csv_matched_fields,
)

FamilyExtractor = Callable[[Worksheet, Dict[str, List[str]]], List[Dict[str, Any]]]
NumberParser = Callable[[Any], float]


class HeaderLocator:
    """엑셀/CSV 파일에서 최적의 시트와 헤더를 감지하는 전담 클래스."""

    def __init__(self, logger: Optional[logging.Logger] = None) -> None:
        self.logger = logger or logging.getLogger(__name__)

    # ------------------------------------------------------------------
    # 공개 API
    # ------------------------------------------------------------------
    def inspect_all_sheets(
        self,
        workbook: openpyxl.Workbook,
        required_keywords: Dict[str, List[str]],
        family_extractor: Optional[FamilyExtractor] = None,
        number_parser: Optional[NumberParser] = None,
    ) -> Optional[Dict[str, Any]]:
        """모든 시트를 검열하여 최적의 시트를 선택한다."""

        weights, thresholds, override_all5 = get_scoring_config()
        forbidden_keywords_map = build_forbidden_keywords_map()

        best_result: Optional[Dict[str, Any]] = None
        sheet_results: List[Dict[str, Any]] = []

        priority_sheet = self._find_priority_sheet(
            workbook,
            required_keywords,
            forbidden_keywords_map,
            family_extractor,
            number_parser,
        )
        if priority_sheet:
            return priority_sheet

        self.logger.info(
            "1순위 시트 없음 - 기존 방식으로 시트 검열 진행"
        )

        sheet_names = sorted(workbook.sheetnames)
        self.logger.info(f"🔍 시트 검열 순서: {sheet_names}")

        for sheet_name in sheet_names:
            try:
                sheet = workbook[sheet_name]
                total_delivery_amount = self._find_max_delivery_amount(sheet)

                sheet_result = self._evaluate_sheet(
                    sheet,
                    sheet_name,
                    required_keywords,
                    forbidden_keywords_map,
                    family_extractor,
                    number_parser,
                )

                if not sheet_result:
                    continue

                # 외부 함수를 사용하여 점수 계산 및 필드 확인
                sheet_result = self._calculate_sheet_scores(
                    sheet_result,
                    weights,
                    total_delivery_amount,
                )

                if total_delivery_amount > 0:
                    delivery_bonus = min(total_delivery_amount / 1_000_000, 50)
                    sheet_result['delivery_bonus'] = delivery_bonus
                    sheet_result['score'] += delivery_bonus
                    self.logger.info(
                        "시트 '%s' 총배달 금액: %s원 (보너스: +%.1f점)",
                        sheet_name,
                        f"{total_delivery_amount:,.0f}",
                        delivery_bonus,
                    )

                sheet_results.append(sheet_result)

                if (
                    sheet_result.get('score', 0) > best_result.get('score', -1)
                    if best_result
                    else True
                ):
                    self.logger.info(
                        "최적 시트 후보 갱신: '%s' (점수 %.2f)",
                        sheet_name,
                        sheet_result.get('score', 0),
                    )
                    best_result = sheet_result

            except Exception as exc:
                self.logger.warning(
                    "시트 '%s' 검열 중 오류: %s",
                    sheet_name,
                    exc,
                )

        if sheet_results:
            delivery_sheets = [r for r in sheet_results if r.get('delivery_amount', 0) > 0]
            if delivery_sheets:
                delivery_sheets.sort(
                    key=lambda r: (r.get('delivery_amount', 0), r.get('sheet_name', '')),
                    reverse=True,
                )
                best_result = delivery_sheets[0]
                self.logger.info(
                    "🎯 총배달 금액 우선 선점: '%s' (총배달 금액: %s원)",
                    best_result['sheet_name'],
                    f"{best_result.get('delivery_amount', 0):,.0f}",
                )
            else:
                best_result = self._select_by_score(
                    sheet_results,
                    thresholds,
                    override_all5,
                )

        if best_result:
            self.logger.info(
                "최적 시트 선택: '%s' (원점수: %.2f, 가중치점: %.2f)",
                best_result['sheet_name'],
                best_result.get('score', 0),
                best_result.get('scoring_points', 0),
            )
        else:
            self.logger.warning("적합한 시트를 찾지 못했습니다.")

        return best_result

    def detect_header_row(self, sheet: Worksheet) -> int:
        """엑셀 시트에서 헤더 행을 감지한다."""
        return detect_header_row(sheet)

    def detect_csv_header_row(
        self,
        df: pd.DataFrame,
        required_keywords: Dict[str, List[str]],
    ) -> int:
        """CSV 파일에서 헤더 행을 감지한다."""
        return detect_csv_header_row(df, required_keywords)

    def map_columns(
        self,
        sheet: Worksheet,
        header_row: int,
        required_keywords: Dict[str, List[str]],
    ) -> Dict[str, int]:
        """필수 컬럼을 헤더 이름 기반으로 매핑한다."""
        return map_columns(sheet, header_row, required_keywords)

    # ------------------------------------------------------------------
    # 내부 유틸리티 (엑셀)
    # ------------------------------------------------------------------
    def _calculate_sheet_scores(
        self,
        sheet_result: Dict[str, Any],
        weights: Dict[str, int],
        total_delivery_amount: float,
    ) -> Dict[str, Any]:
        """시트 점수를 계산하고 필드 존재 여부를 확인한다."""
        headers_lower = [str(h).lower() for h in sheet_result["headers"]]

        def any_in(keys: List[str]) -> bool:
            return any(
                any(key in header for header in headers_lower) for key in keys
            )

        has_bn = any_in(['사업자', '등록번호', '공급받는자사업자', '공급받는자 사업자'])

        representative_scores = [
            score_representative_header(header)
            for header in headers_lower
        ]

        strong_rep = any(score >= 60 for score in representative_scores)
        medium_rep = any(40 <= score < 60 for score in representative_scores)
        fallback_rep = any(20 <= score < 40 for score in representative_scores)
        registration_rep = any(
            ('등록자' in header or '작성자' in header or '입력자' in header)
            for header in headers_lower
        )

        if strong_rep:
            representative_points = weights.get('representative', 10)
        elif medium_rep:
            representative_points = max(int(weights.get('representative', 10) * 0.6), 4)
        elif fallback_rep:
            representative_points = max(int(weights.get('representative', 10) * 0.4), 3)
        elif registration_rep:
            representative_points = max(int(weights.get('representative', 10) * 0.2), 2)
        else:
            representative_points = 0

        has_rep = representative_points > 0
        has_addr = any_in(['주소', '소재지', '사업장', '공급받는자주소', '공급받는자 주소'])
        has_email = any_in(['이메일', 'email', '메일', '공급받는자이메일', '공급받는자 이메일'])
        has_store = any_in(['가맹점', '상호', '상호명', '매장', '점포', '업체', '가게', '공급받는자상호', '공급받는자 상호', '매장명', '점포명'])

        found5 = sum([has_bn, has_rep, has_addr, has_email, has_store])

        sheet_score_pts = (
            (weights.get('business_number', 30) if has_bn else 0)
            + representative_points
            + (weights.get('address', 30) if has_addr else 0)
            + (weights.get('email', 20) if has_email else 0)
            + (weights.get('store_name', 10) if has_store else 0)
        )

        sheet_result['core_fields_found'] = found5
        sheet_result['scoring_points'] = sheet_score_pts
        sheet_result['delivery_amount'] = total_delivery_amount

        return sheet_result

    def _find_priority_sheet(
        self,
        workbook: openpyxl.Workbook,
        required_keywords: Dict[str, List[str]],
        forbidden_keywords_map: Dict[str, List[str]],
        family_extractor: Optional[FamilyExtractor],
        number_parser: Optional[NumberParser],
    ) -> Optional[Dict[str, Any]]:
        self.logger.info("🚀 1순위 시트 Fast-Path 검색 시작")

        best_sheet: Optional[Dict[str, Any]] = None
        max_dad_amount = 0.0

        sheet_names = sorted(workbook.sheetnames)
        self.logger.info(f"🔍 시트 검색 순서: {sheet_names}")

        for sheet_name in sheet_names:
            try:
                sheet = workbook[sheet_name]
                sheet_result = self._evaluate_sheet(
                    sheet,
                    sheet_name,
                    required_keywords,
                    forbidden_keywords_map,
                    family_extractor,
                    number_parser,
                )

                if not sheet_result:
                    continue

                matched_fields = sheet_result.get('matched_fields', 0)
                families = sheet_result.get('families', [])

                # 대혁명 2단계: 현명한 왕의 법률 - 더 유연한 조건
                if matched_fields >= 5 or (matched_fields >= 4 and families):
                    max_dad_with_mom = self._get_max_dad_with_mom_same_row(
                        sheet_result,
                        number_parser,
                    )

                    # 대혁명 2단계: 보안관 임명 - 압도적인 1순위 시트 발견 시 즉시 종료
                    if max_dad_with_mom >= 1_000_000:
                        best_sheet = sheet_result
                        best_sheet['max_dad_amount'] = max_dad_with_mom
                        best_sheet['priority'] = '1순위'
                        
                        self.logger.info(
                            "👑 왕을 발견했습니다: '%s' (아빠값: %s원). 즉시 수색을 종료합니다.",
                            sheet_name,
                            f"{max_dad_with_mom:,.0f}",
                        )
                        return best_sheet

                    if (
                        max_dad_with_mom > max_dad_amount
                        or (
                            max_dad_with_mom == max_dad_amount
                            and (
                                best_sheet is None
                                or sheet_name < best_sheet['sheet_name']
                            )
                        )
                    ):
                        max_dad_amount = max_dad_with_mom
                        best_sheet = sheet_result
                        best_sheet['max_dad_amount'] = max_dad_amount
                        best_sheet['priority'] = '1순위'

                        self.logger.info(
                            "🎯 1순위 시트 후보 발견: '%s' (아빠값: %s원)",
                            sheet_name,
                            f"{max_dad_amount:,.0f}",
                        )

            except Exception as exc:
                self.logger.warning(
                    "시트 '%s' 1순위 검색 중 오류: %s",
                    sheet_name,
                    exc,
                )

        if best_sheet:
            self.logger.info(
                "🏆 1순위 시트 최종 선택: '%s' (아빠값: %s원)",
                best_sheet['sheet_name'],
                f"{max_dad_amount:,.0f}",
            )
        else:
            self.logger.info("❌ 1순위 시트를 찾지 못했습니다.")

        return best_sheet

    def _evaluate_sheet(
        self,
        sheet: Worksheet,
        sheet_name: str,
        required_keywords: Dict[str, List[str]],
        forbidden_keywords_map: Dict[str, List[str]],
        family_extractor: Optional[FamilyExtractor],
        number_parser: Optional[NumberParser],
    ) -> Optional[Dict[str, Any]]:
        try:
            actual_max_row, actual_max_col = find_actual_data_range(sheet)
            max_row = actual_max_row
            max_col = min(actual_max_col, 50)

            if max_row < 2 or max_col < 5:
                return None

            header_candidates: List[Dict[str, Any]] = []
            scan_rows = min(1000, max_row)

            for row in range(1, scan_rows + 1):
                data_density = calculate_data_density(sheet, row, max_col)
                matched_fields = count_matched_fields(
                    sheet,
                    row,
                    max_col,
                    required_keywords,
                    forbidden_keywords_map,
                )

                field_match_score = (matched_fields / 5) * 0.8
                density_score = data_density * 0.2
                header_score = field_match_score + density_score

                if matched_fields >= 3:
                    non_empty_cells = 0
                    for col in range(1, max_col + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value is not None and str(cell_value).strip():
                            non_empty_cells += 1

                    if non_empty_cells >= 5:
                        header_candidates.append(
                            {
                                'row': row,
                                'data_density': data_density,
                                'matched_fields': matched_fields,
                                'header_score': header_score,
                                'non_empty_cells': non_empty_cells,
                            }
                        )
                        self.logger.debug(
                            "지능앱 헤더 후보: 행 %d, 점수 %.3f, 매칭 %d개, 비어있지 않은 셀 %d개",
                            row,
                            header_score,
                            matched_fields,
                            non_empty_cells,
                        )

            if not header_candidates:
                self.logger.warning(
                    "지능앱 헤더 감지: 시트 '%s'에서 적합한 헤더를 찾지 못했습니다.",
                    sheet_name,
                )
                return None

            header_candidates.sort(
                key=lambda x: (
                    x['matched_fields'],
                    x['header_score'],
                    -x['row'],
                    x['row'],
                ),
                reverse=True,
            )

            best_header = header_candidates[0]
            header_row = best_header['row']
            self.logger.info(
                "지능앱 헤더 감지: 시트 '%s'에서 최적 헤더 선택 - 행 %d (점수: %.3f, 매칭: %d개)",
                sheet_name,
                header_row,
                best_header['header_score'],
                best_header['matched_fields'],
            )

            headers: List[str] = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")

            data: List[List[Any]] = []
            data_start_row = header_row + 1
            for row in range(data_start_row, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(cell_value)
                data.append(row_data)

            data_quality_score = evaluate_data_quality(data, headers)
            final_score = best_header['header_score'] * 0.7 + data_quality_score * 0.3

            families: List[Dict[str, Any]] = []
            if family_extractor:
                try:
                    families = family_extractor(sheet, required_keywords)
                except Exception as exc:  # pragma: no cover - 안전 장치
                    self.logger.warning("가족 정보 추출 중 오류: %s", exc)

            return {
                'sheet_name': sheet_name,
                'header_row': header_row,
                'data_start_row': data_start_row,
                'headers': headers,
                'data': data,
                'score': final_score,
                'matched_fields': best_header['matched_fields'],
                'data_quality': data_quality_score,
                'total_rows': len(data),
                'actual_max_row': actual_max_row,
                'actual_max_col': actual_max_col,
                'families': families,
            }

        except Exception as exc:
            self.logger.error("시트 '%s' 평가 오류: %s", sheet_name, exc)
            return None

    def get_actual_data_range(self, sheet: Worksheet) -> Tuple[int, int]:
        """공개용: 실제 데이터 범위를 반환한다."""
        return get_actual_data_range(sheet)

    def _find_max_delivery_amount(self, sheet: Worksheet) -> float:
        max_amount = 0.0
        delivery_keywords = ['총배달', '배달금액', '총금액', '합계', '총합', '배달요금', '총배달요금', '총배달금액']

        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_value = str(cell.value).lower().strip()

                    if any(keyword in cell_value for keyword in delivery_keywords):
                        col_idx = cell.column
                        for data_row in sheet.iter_rows(min_row=cell.row + 1, min_col=col_idx, max_col=col_idx):
                            for data_cell in data_row:
                                if (
                                    isinstance(data_cell.value, (int, float))
                                    and data_cell.value > max_amount
                                ):
                                    max_amount = float(data_cell.value)
                        break
        except Exception as exc:
            self.logger.warning("총배달 금액 분석 중 오류: %s", exc)

        return max_amount

    def _get_max_dad_with_mom_same_row(
        self,
        sheet_result: Dict[str, Any],
        number_parser: Optional[NumberParser],
    ) -> float:
        try:
            families = sheet_result.get('families', [])
            if not families:
                return 0.0

            max_dad_amount = 0.0
            converter = number_parser or default_number_parser

            for family in families:
                dad_amount = converter(family.get('공급가액', 0))
                mom_amount = converter(family.get('부가세', 0))

                if dad_amount > 0 and mom_amount > 0:
                    if dad_amount > max_dad_amount:
                        max_dad_amount = dad_amount

            self.logger.debug(
                "🔍 같은 행 아빠값 최대값: %s원",
                f"{max_dad_amount:,.0f}",
            )
            return max_dad_amount
        except Exception as exc:
            self.logger.warning("같은 행 아빠값 최대값 계산 중 오류: %s", exc)
            return 0.0


    # ------------------------------------------------------------------
    # 설정/빌더
    # ------------------------------------------------------------------
    def _select_by_score(
        self,
        sheet_results: List[Dict[str, Any]],
        thresholds: Dict[str, int],
        override_all5: bool,
    ) -> Optional[Dict[str, Any]]:
        if not sheet_results:
            return None

        if override_all5:
            all5 = [r for r in sheet_results if r.get('core_fields_found', 0) >= 5]
            if all5:
                all5.sort(key=lambda r: (
                    r.get('scoring_points', 0),
                    r.get('total_rows', 0),
                    -r.get('header_row', 0),
                    r.get('sheet_name', ''),
                ), reverse=True)
                return all5[0]

            passing = [r for r in sheet_results if r.get('scoring_points', 0) >= thresholds.get('pass', 80)]
            if passing:
                passing.sort(key=lambda r: (
                    r.get('scoring_points', 0),
                    r.get('total_rows', 0),
                    -r.get('header_row', 0),
                    r.get('sheet_name', ''),
                ), reverse=True)
                return passing[0]

            candidates = [r for r in sheet_results if r.get('scoring_points', 0) >= thresholds.get('candidate', 70)]
            if candidates:
                candidates.sort(key=lambda r: (
                    r.get('scoring_points', 0),
                    r.get('total_rows', 0),
                    -r.get('header_row', 0),
                    r.get('sheet_name', ''),
                ), reverse=True)
                return candidates[0]

        sheet_results.sort(key=lambda r: (r.get('score', 0), r.get('sheet_name', '')), reverse=True)
        return sheet_results[0]

    # ------------------------------------------------------------------
    # 내부 공용 유틸리티
    # ------------------------------------------------------------------


