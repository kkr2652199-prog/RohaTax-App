"""
헤더 감지 시스템 모듈
Excel/CSV 파일의 헤더 행을 지능적으로 감지하는 시스템
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, Tuple
import logging

logger = logging.getLogger(__name__)


class HeaderDetector:
    """지능형 헤더 감지 시스템"""
    
    def __init__(self):
        self.logger = logger
    
    def detect_header_row(self, sheet) -> int:
        """
        헤더 행 자동 감지
        지능앱 기술: 제목/부제목 행을 건너뛰고 실제 헤더 행 찾기
        
        Returns:
            int: 헤더 행 번호 (1부터 시작)
        """
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # 🎯 제목/부제목 행 감지 및 건너뛰기
        title_rows = self._detect_title_rows(sheet, max_col)
        self.logger.info(f"🔍 제목/부제목 행 감지: {title_rows}")
        
        # 상위 1000행에서 헤더 후보 검색 (제목 행 제외)
        header_candidates = []
        
        for row in range(1, min(1001, max_row + 1)):
            # 제목/부제목 행은 건너뛰기
            if row in title_rows:
                continue
                
            # 해당 행의 데이터 밀도 계산
            data_density = self._calculate_data_density(sheet, row, max_col)
            header_candidates.append((row, data_density))
        
        if not header_candidates:
            # 헤더 후보가 없으면 기본값 사용
            self.logger.warning("헤더 후보가 없습니다. 기본 헤더 행 사용")
            return 1
        
        # 데이터 밀도가 가장 높은 행을 헤더로 선택 (결정적 정렬)
        header_candidates.sort(key=lambda x: (x[1], x[0]), reverse=True)  # 밀도, 행번호 순
        best_header_row = header_candidates[0][0]
        
        self.logger.info(f"헤더 후보 분석: {header_candidates[:5]}...")  # 처음 5개만 로그
        self.logger.info(f"선택된 헤더 행: {best_header_row}")
        
        return best_header_row
    
    def _detect_title_rows(self, sheet, max_col: int) -> List[int]:
        """
        제목/부제목 행 감지
        지능앱 기술: 텍스트 패턴 분석으로 제목 행 식별
        """
        title_rows = []
        
        # 상위 10행에서 제목 패턴 검색
        for row in range(1, min(11, sheet.max_row + 1)):
            text_pattern = self._analyze_text_pattern(sheet, row, max_col)
            
            if self._is_title_pattern(text_pattern):
                title_rows.append(row)
                self.logger.info(f"제목 행 감지: {row}행 - {text_pattern}")
        
        return title_rows
    
    def _analyze_text_pattern(self, sheet, row: int, max_col: int) -> Dict[str, Any]:
        """텍스트 패턴 분석"""
        pattern = {
            'total_cells': 0,
            'text_cells': 0,
            'numeric_cells': 0,
            'empty_cells': 0,
            'merged_cells': 0,
            'text_length': 0
        }
        
        try:
            for col in range(1, min(max_col + 1, 50)):  # 최대 50열까지만 검사
                cell = sheet.cell(row=row, column=col)
                pattern['total_cells'] += 1
                
                if cell.value is None:
                    pattern['empty_cells'] += 1
                elif isinstance(cell.value, (int, float)):
                    pattern['numeric_cells'] += 1
                else:
                    pattern['text_cells'] += 1
                    pattern['text_length'] += len(str(cell.value))
                
                # 병합된 셀 확인
                if sheet.cell(row=row, column=col).coordinate in sheet.merged_cells:
                    pattern['merged_cells'] += 1
                    
        except Exception as e:
            self.logger.warning(f"텍스트 패턴 분석 중 오류: {str(e)}")
        
        return pattern
    
    def _is_title_pattern(self, pattern: Dict[str, Any]) -> bool:
        """제목 패턴인지 판단"""
        # 제목 패턴 조건:
        # 1. 텍스트 셀이 많고 숫자 셀이 적음
        # 2. 병합된 셀이 있음
        # 3. 전체 텍스트 길이가 길음
        
        if pattern['total_cells'] == 0:
            return False
        
        text_ratio = pattern['text_cells'] / pattern['total_cells']
        numeric_ratio = pattern['numeric_cells'] / pattern['total_cells']
        avg_text_length = pattern['text_length'] / max(pattern['text_cells'], 1)
        
        # 제목 패턴 판단 기준
        is_title = (
            text_ratio > 0.7 and  # 텍스트 비율이 높음
            numeric_ratio < 0.3 and  # 숫자 비율이 낮음
            avg_text_length > 5 and  # 평균 텍스트 길이가 길음
            pattern['merged_cells'] > 0  # 병합된 셀이 있음
        )
        
        return is_title
    
    def _calculate_data_density(self, sheet, row: int, max_col: int) -> float:
        """데이터 밀도 계산 (헤더 후보 평가)"""
        try:
            total_cells = 0
            filled_cells = 0
            
            # 해당 행의 모든 셀 검사
            for col in range(1, min(max_col + 1, 50)):  # 최대 50열까지만 검사
                cell = sheet.cell(row=row, column=col)
                total_cells += 1
                
                if cell.value is not None and str(cell.value).strip():
                    filled_cells += 1
            
            # 데이터 밀도 계산 (0.0 ~ 1.0)
            density = filled_cells / max(total_cells, 1)
            
            return density
            
        except Exception as e:
            self.logger.warning(f"데이터 밀도 계산 중 오류: {str(e)}")
            return 0.0
    
    def find_header_row(self, sheet) -> Optional[int]:
        """유연한 헤더 감지 시스템 (동적 컬럼 수 + 5형제 키워드 기반)"""
        # 5형제 키워드 정의
        header_keywords = ['사업자번호', '사업자등록번호', '상호명', '상호', '업체명', '가맹점명', 
                          '대표자', '대표자명', '주소', '사업장주소', '이메일', 'email']
        
        best_header_row = None
        best_score = 0
        
        for row_num in range(1, min(15, sheet.max_row + 1)):
            row_values = [str(cell.value).strip() for cell in sheet[row_num] if cell.value]
            
            # 최소 5개 컬럼이 있어야 헤더로 인정 (5형제 최소 요구사항)
            if len(row_values) >= 5:
                # 5형제 키워드 카운트
                header_count = 0
                for cell_value in row_values:
                    if any(keyword in cell_value for keyword in header_keywords):
                        header_count += 1
                
                # 5형제 키워드가 2개 이상 포함된 행을 헤더로 인정
                if header_count >= 2:
                    # 컬럼 수와 키워드 수를 종합한 점수 계산
                    score = header_count * 10 + len(row_values)
                    
                    if score > best_score:
                        best_score = score
                        best_header_row = row_num
        
        return best_header_row
    
    def calculate_header_data_quality(self, sheet, header_row: int) -> float:
        """헤더 데이터 품질 평가"""
        try:
            quality_score = 0.0
            
            # 헤더 행의 데이터 품질 평가
            for col in range(1, min(sheet.max_column + 1, 50)):
                cell = sheet.cell(row=header_row, column=col)
                if cell.value and str(cell.value).strip():
                    quality_score += 1.0
            
            # 정규화 (0.0 ~ 1.0)
            max_cols = min(sheet.max_column, 50)
            normalized_score = quality_score / max(max_cols, 1)
            
            return normalized_score
            
        except Exception as e:
            self.logger.warning(f"헤더 데이터 품질 평가 중 오류: {str(e)}")
            return 0.0
    
    def detect_csv_header_row(self, df: pd.DataFrame) -> int:
        """CSV 파일의 헤더 행 감지"""
        best_header_row = 0
        best_score = 0
        
        # 상위 10행에서 헤더 후보 검색
        for row_idx in range(min(10, len(df))):
            score = self._calculate_csv_data_density(df, row_idx)
            
            if score > best_score:
                best_score = score
                best_header_row = row_idx
        
        return best_header_row
    
    def _calculate_csv_data_density(self, df: pd.DataFrame, row_idx: int) -> float:
        """CSV 데이터 밀도 계산"""
        try:
            row_data = df.iloc[row_idx]
            total_cells = len(row_data)
            filled_cells = 0
            
            for value in row_data:
                if pd.notna(value) and str(value).strip():
                    filled_cells += 1
            
            return filled_cells / max(total_cells, 1)
            
        except Exception as e:
            self.logger.warning(f"CSV 데이터 밀도 계산 중 오류: {str(e)}")
            return 0.0
    
    def quick_header_scan(self, sheet, max_rows: int = 5) -> Optional[Dict]:
        """빠른 헤더 스캔 (성능 최적화)"""
        def _norm(s: str) -> str:
            return str(s).strip().replace('\n', '').replace(' ', '').lower()
        
        try:
            # 상위 몇 행만 빠르게 스캔
            for row_num in range(1, min(max_rows + 1, sheet.max_row + 1)):
                row_values = [str(cell.value).strip() for cell in sheet[row_num] if cell.value]
                
                if len(row_values) >= 3:  # 최소 3개 컬럼
                    # 간단한 키워드 체크
                    keywords_found = 0
                    for value in row_values:
                        norm_value = _norm(value)
                        if any(kw in norm_value for kw in ['사업자', '상호', '대표자', '주소', '이메일']):
                            keywords_found += 1
                    
                    if keywords_found >= 2:  # 2개 이상 키워드 발견
                        return {
                            'header_row': row_num,
                            'confidence': keywords_found / len(row_values),
                            'columns': len(row_values)
                        }
            
            return None
            
        except Exception as e:
            self.logger.warning(f"빠른 헤더 스캔 중 오류: {str(e)}")
            return None


